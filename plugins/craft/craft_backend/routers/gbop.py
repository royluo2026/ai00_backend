"""
backend/routers/gbop.py
──────────────────────────
GBOP 标准工序库 V3 API（树形结构 + 独立工艺/操作实体 + entry_links）

端点：
  GET/POST       /api/gbop/versions                         版本列表/创建
  PATCH          /api/gbop/versions/{gid}                   更新版本
  POST           /api/gbop/versions/{gid}/freeze            冻结
  POST           /api/gbop/version-families/{fgid}/archive  归档族
  DELETE         /api/gbop/version-families/{fgid}/archive  解除归档
  GET            /api/gbop/versions/{gid}/entries           获取全部节点（含 links）
  POST           /api/gbop/entries                          创建节点
  PATCH          /api/gbop/entries/{gid}                    更新节点
  DELETE         /api/gbop/entries/{gid}                    删除节点
  GET            /api/gbop/versions/{gid}/processes         获取版本下所有工艺卡片
  POST           /api/gbop/processes                        创建工艺卡片（一键模式）
  PATCH          /api/gbop/processes/{gid}                  更新工艺卡片
  DELETE         /api/gbop/processes/{gid}                  删除工艺卡片
  GET            /api/gbop/versions/{gid}/operations        获取版本下所有操作卡片
  POST           /api/gbop/operations                       创建操作卡片（一键模式）
  PATCH          /api/gbop/operations/{gid}                 更新操作卡片
  DELETE         /api/gbop/operations/{gid}                 删除操作卡片
  POST           /api/gbop/entry-links                      手动创建挂载链接
  DELETE         /api/gbop/entry-links/{gid}                删除挂载链接
  GET            /api/gbop/entries/{gid}/links              获取某 entry 的所有挂载
  POST           /api/gbop/versions/{gid}/import-vpps-parts 从 vpps_parts 导入 L1-3
  POST           /api/gbop/versions/{gid}/import-entries    Excel 批量导入
  POST           /api/gbop/versions/{gid}/import-tc-excel  TC 两 Sheet Excel 导入工序/操作+绑定
  POST           /api/gbop/versions/{gid}/fork              Fork 版本
"""
import json
from typing import Dict, List, Literal, Optional

import io

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from pydantic import BaseModel

from ..data.connection import get_conn
from backend.platform_sdk.auth import require_role
from backend.platform_sdk.ids import next_gid

router = APIRouter(prefix="/api/gbop", tags=["gbop"])

_WRITE = require_role("super_admin", "team_admin", "project_admin", "knowledge_admin", "member")
_READ = require_role("super_admin", "team_admin", "project_admin",
                     "rule_admin", "knowledge_admin", "member")


# ── Pydantic Models ──────────────────────────────────────────────

class CreateVersionBody(BaseModel):
    name: str = ''
    version_family_gid: Optional[str] = None  # None = 新建族
    vehicle_model: str = ''

class UpdateVersionBody(BaseModel):
    name: Optional[str] = None
    vehicle_model: Optional[str] = None
    status: Optional[str] = None

class CreateEntryBody(BaseModel):
    version_gid: str
    parent_gid: Optional[str] = None
    node_type: str = 'process'
    seq_no: float = 0
    vpps: Optional[str] = None
    vpps_desc: str = ''
    vpps_attr: str = ''
    importance: str = ''
    torque_importance: str = ''
    vehicle_model: str = ''
    parent_vpps: str = ''
    meta: dict = {}
    vpps_part: str = ''
    part_feed: bool = False

class UpdateEntryBody(BaseModel):
    parent_gid: Optional[str] = None
    node_type: Optional[str] = None
    seq_no: Optional[float] = None
    vpps: Optional[str] = None
    vpps_desc: Optional[str] = None
    vpps_attr: Optional[str] = None
    importance: Optional[str] = None
    torque_importance: Optional[str] = None
    vehicle_model: Optional[str] = None
    parent_vpps: Optional[str] = None
    status: Optional[str] = None
    meta: Optional[dict] = None
    vpps_part: Optional[str] = None
    part_feed: Optional[bool] = None

class CreateProcessBody(BaseModel):
    version_gid: str
    parent_entry_gid: Optional[str] = None  # 自动创建 entry 的父节点
    vpps: Optional[str] = None
    vpps_desc: str = ''
    op_code: str = ''
    op_name: str = ''
    standard_time: Optional[float] = None
    description: str = ''
    steps: list = []
    required_tools: list = []
    parameters: dict = {}
    importance: str = ''
    torque_importance: str = ''
    vehicle_model: str = ''
    meta: dict = {}
    seq_no: float = 0
    vpps_part: str = ''
    part_feed: bool = False

class UpdateProcessBody(BaseModel):
    vpps: Optional[str] = None
    vpps_desc: Optional[str] = None
    op_code: Optional[str] = None
    op_name: Optional[str] = None
    standard_time: Optional[float] = None
    description: Optional[str] = None
    steps: Optional[list] = None
    required_tools: Optional[list] = None
    parameters: Optional[dict] = None
    importance: Optional[str] = None
    torque_importance: Optional[str] = None
    vehicle_model: Optional[str] = None
    status: Optional[str] = None
    meta: Optional[dict] = None
    vpps_part: Optional[str] = None
    part_feed: Optional[bool] = None

class CreateOperationBody(BaseModel):
    version_gid: str
    process_gid: Optional[str] = None
    parent_entry_gid: Optional[str] = None
    vpps: Optional[str] = None
    vpps_desc: str = ''
    op_code: str = ''
    op_name: str = ''
    standard_time: Optional[float] = None
    description: str = ''
    steps: list = []
    required_tools: list = []
    parameters: dict = {}
    importance: str = ''
    torque_importance: str = ''
    vehicle_model: str = ''
    meta: dict = {}
    seq_no: float = 0
    vpps_part: str = ''
    part_feed: bool = False

class UpdateOperationBody(BaseModel):
    process_gid: Optional[str] = None
    vpps: Optional[str] = None
    vpps_desc: Optional[str] = None
    op_code: Optional[str] = None
    op_name: Optional[str] = None
    standard_time: Optional[float] = None
    description: Optional[str] = None
    steps: Optional[list] = None
    required_tools: Optional[list] = None
    parameters: Optional[dict] = None
    importance: Optional[str] = None
    torque_importance: Optional[str] = None
    vehicle_model: Optional[str] = None
    status: Optional[str] = None
    meta: Optional[dict] = None
    vpps_part: Optional[str] = None
    part_feed: Optional[bool] = None

class CreateEntryLinkBody(BaseModel):
    entry_gid: str
    link_type: str  # 'gbop_process' | 'gbop_operation'
    ref_gid: str
    is_primary: bool = False

class ImportVppsPartsBody(BaseModel):
    levels: List[int] = [1, 2, 3]  # 导入哪些层级

class ImportEntriesBody(BaseModel):
    entries: List[dict]  # 批量导入条目

class ForkBody(BaseModel):
    target_name: Optional[str] = None
    target_version_family_gid: Optional[str] = None
    change_note: Optional[str] = None
    include_node_types: Optional[List[str]] = None


# ── Helpers ──────────────────────────────────────────────────────

_VER_COLS = ("gid,name,version_family_gid,status,frozen_at,archived_at,"
             "vehicle_model,team_id,created_by,created_at,updated_at")
_VER_KEYS = ['gid', 'name', 'version_family_gid', 'status', 'frozen_at', 'archived_at',
             'vehicle_model', 'team_id', 'created_by', 'created_at', 'updated_at']

_ENTRY_COLS = ("gid,version_gid,parent_gid,level,node_type,seq_no,"
               "vpps,vpps_desc,vpps_attr,importance,torque_importance,"
               "vehicle_model,parent_vpps,status,sort_order,"
               "meta,team_id,created_by,created_at,updated_at,"
               "vpps_part,part_feed")

_PROC_COLS = ("gid,version_gid,vpps,vpps_desc,op_code,op_name,standard_time,"
              "description,steps,required_tools,parameters,"
              "importance,torque_importance,vehicle_model,status,meta,"
              "created_by,created_at,updated_at,"
              "vpps_part,part_feed")

_OP_COLS = ("gid,version_gid,process_gid,vpps,vpps_desc,op_code,op_name,standard_time,"
            "description,steps,required_tools,parameters,"
            "importance,torque_importance,vehicle_model,status,meta,"
            "created_by,created_at,updated_at,"
            "vpps_part,part_feed")

_LINK_COLS = "gid,entry_gid,link_type,ref_gid,is_primary,created_at,created_by"


def _ver_row(row) -> dict:
    if not row:
        return None
    return {k: (str(row[k]) if k in ('created_at', 'updated_at', 'frozen_at', 'archived_at') and row[k] else row[k])
            for k in _VER_KEYS}


def _entry_row(row) -> dict:
    if not row:
        return None
    d = dict(row)
    for k in ('created_at', 'updated_at'):
        if d.get(k):
            d[k] = str(d[k])
    return d


def _entity_row(row) -> dict:
    if not row:
        return None
    d = dict(row)
    for k in ('created_at', 'updated_at'):
        if d.get(k):
            d[k] = str(d[k])
    return d


def _link_row(row) -> dict:
    if not row:
        return None
    d = dict(row)
    if d.get('created_at'):
        d['created_at'] = str(d['created_at'])
    return d


def _check_frozen(cur, version_gid: str):
    """若版本已冻结则抛 403"""
    cur.execute("SELECT frozen_at FROM workmanship_tpl_gbop_versions WHERE gid=%s", (version_gid,))
    row = cur.fetchone()
    if not row:
        raise HTTPException(404, "版本不存在")
    if row['frozen_at']:
        raise HTTPException(403, "版本已冻结，不允许修改")


def _calc_level(cur, parent_gid: Optional[str]) -> int:
    """计算节点 level = parent.level + 1"""
    if not parent_gid:
        return 0
    cur.execute("SELECT level FROM workmanship_tpl_gbop_entries WHERE gid=%s", (parent_gid,))
    row = cur.fetchone()
    if not row:
        return 0
    return row['level'] + 1


# ── Version CRUD ─────────────────────────────────────────────────

@router.get("/versions")
def list_versions(
    include_archived: bool = False,
    current_user: dict = Depends(_READ)
):
    with get_conn() as conn:
        with conn.cursor() as cur:
            conditions = []
            params = []
            if not include_archived:
                conditions.append("archived_at IS NULL")
            where = ("WHERE " + " AND ".join(conditions)) if conditions else ""
            cur.execute(
                f"SELECT {_VER_COLS} FROM workmanship_tpl_gbop_versions "
                f"{where} ORDER BY version_family_gid, created_at",
                params
            )
            rows = cur.fetchall()
    return {"data": [_ver_row(r) for r in rows]}


@router.post("/versions", status_code=201)
def create_version(body: CreateVersionBody, current_user: dict = Depends(_WRITE)):
    gid = str(next_gid())
    family_gid = body.version_family_gid or gid
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute(
                f"INSERT INTO workmanship_tpl_gbop_versions "
                f"(gid, name, version_family_gid, vehicle_model, team_id, created_by) "
                f"VALUES (%s, %s, %s, %s, %s, %s)",
                (gid, body.name, family_gid, body.vehicle_model,
                 current_user.get("team_id"), current_user["gid"])
            )
            cur.execute(f"SELECT {_VER_COLS} FROM workmanship_tpl_gbop_versions WHERE gid=%s", (gid,))
            row = cur.fetchone()
            conn.commit()
    return {"data": _ver_row(row)}


@router.patch("/versions/{gid}")
def update_version(gid: str, body: UpdateVersionBody, current_user: dict = Depends(_WRITE)):
    data = body.model_dump(exclude_unset=True)
    if not data:
        raise HTTPException(400, "无更新字段")
    set_parts, vals = [], []
    for k, v in data.items():
        set_parts.append(f"{k}=%s")
        vals.append(v)
    set_parts.append("updated_at=NOW()")
    vals.append(gid)
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute(
                f"UPDATE workmanship_tpl_gbop_versions SET {', '.join(set_parts)} "
                f"WHERE gid=%s",
                vals
            )
            if cur.rowcount == 0:
                raise HTTPException(404, "版本不存在")
            cur.execute(f"SELECT {_VER_COLS} FROM workmanship_tpl_gbop_versions WHERE gid=%s", (gid,))
            row = cur.fetchone()
            conn.commit()
    return {"data": _ver_row(row)}


@router.post("/versions/{gid}/freeze")
def freeze_version(gid: str, _u=Depends(_WRITE)):
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute(
                f"UPDATE workmanship_tpl_gbop_versions SET status='frozen', frozen_at=NOW(), updated_at=NOW() "
                f"WHERE gid=%s AND frozen_at IS NULL",
                (gid,)
            )
            if cur.rowcount == 0:
                raise HTTPException(400, "版本不存在或已冻结")
            cur.execute(f"SELECT {_VER_COLS} FROM workmanship_tpl_gbop_versions WHERE gid=%s", (gid,))
            row = cur.fetchone()
            conn.commit()
    return {"data": _ver_row(row)}


@router.post("/version-families/{family_gid}/archive")
def archive_family(family_gid: str, _u=Depends(_WRITE)):
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "UPDATE workmanship_tpl_gbop_versions SET archived_at=NOW(), updated_at=NOW() "
                "WHERE version_family_gid=%s",
                (family_gid,)
            )
            count = cur.rowcount
            conn.commit()
    return {"data": {"archived_count": count}}


@router.delete("/version-families/{family_gid}/archive")
def unarchive_family(family_gid: str, _u=Depends(_WRITE)):
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "UPDATE workmanship_tpl_gbop_versions SET archived_at=NULL, updated_at=NOW() "
                "WHERE version_family_gid=%s",
                (family_gid,)
            )
            count = cur.rowcount
            conn.commit()
    return {"data": {"unarchived_count": count}}


# ── Entry CRUD ───────────────────────────────────────────────────

@router.get("/versions/{version_gid}/entries")
def list_entries(version_gid: str, current_user: dict = Depends(_READ)):
    """获取全部节点，每条 entry 附带 links 数组"""
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute(
                f"SELECT {_ENTRY_COLS} FROM workmanship_tpl_gbop_entries "
                f"WHERE version_gid=%s ORDER BY seq_no, created_at",
                (version_gid,)
            )
            entries = cur.fetchall()

            # 批量获取该版本所有 entry_links
            cur.execute(
                f"SELECT {_LINK_COLS} FROM workmanship_tpl_gbop_entry_links "
                f"WHERE entry_gid IN (SELECT gid FROM workmanship_tpl_gbop_entries WHERE version_gid=%s)",
                (version_gid,)
            )
            links = cur.fetchall()

    # 按 entry_gid 分组 links
    links_by_entry = {}
    for lk in links:
        eg = lk['entry_gid']
        if eg not in links_by_entry:
            links_by_entry[eg] = []
        links_by_entry[eg].append(_link_row(lk))

    result = []
    for e in entries:
        d = _entry_row(e)
        d['links'] = links_by_entry.get(e['gid'], [])
        result.append(d)

    return {"data": result}


@router.post("/entries", status_code=201)
def create_entry(body: CreateEntryBody, current_user: dict = Depends(_WRITE)):
    gid = str(next_gid())
    with get_conn() as conn:
        with conn.cursor() as cur:
            _check_frozen(cur, body.version_gid)
            level = _calc_level(cur, body.parent_gid)
            cur.execute(
                "INSERT INTO workmanship_tpl_gbop_entries "
                "(gid, version_gid, parent_gid, level, node_type, seq_no, "
                " vpps, vpps_desc, vpps_attr, importance, torque_importance, "
                " vehicle_model, parent_vpps, meta, "
                " team_id, created_by, vpps_part, part_feed) "
                "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,"
                "        %s,%s,%s,%s,%s) ",
                (gid, body.version_gid, body.parent_gid, level, body.node_type, body.seq_no,
                 body.vpps, body.vpps_desc, body.vpps_attr, body.importance,
                 body.torque_importance, body.vehicle_model, body.parent_vpps,
                 json.dumps(body.meta, ensure_ascii=False),
                 current_user.get("team_id"), current_user["gid"],
                 body.vpps_part, body.part_feed)
            )
            cur.execute(f"SELECT {_ENTRY_COLS} FROM workmanship_tpl_gbop_entries WHERE gid=%s", (gid,))
            row = cur.fetchone()
            conn.commit()
    return {"data": _entry_row(row)}


@router.patch("/entries/{gid}")
def update_entry(gid: str, body: UpdateEntryBody, current_user: dict = Depends(_WRITE)):
    data = body.model_dump(exclude_unset=True)
    if not data:
        raise HTTPException(400, "无更新字段")

    with get_conn() as conn:
        with conn.cursor() as cur:
            # 查询所属版本，检查冻结
            cur.execute("SELECT version_gid FROM workmanship_tpl_gbop_entries WHERE gid=%s", (gid,))
            entry_row = cur.fetchone()
            if not entry_row:
                raise HTTPException(404, "节点不存在")
            _check_frozen(cur, entry_row['version_gid'])

            # 如果修改了 parent_gid，重算 level
            if 'parent_gid' in data:
                data['level'] = _calc_level(cur, data['parent_gid'])

            json_fields = {'meta'}
            set_parts, vals = [], []
            for k, v in data.items():
                if k in json_fields:
                    set_parts.append(f"{k}=%s")
                    vals.append(json.dumps(v, ensure_ascii=False) if v is not None else '{}')
                else:
                    set_parts.append(f"{k}=%s")
                    vals.append(v)
            set_parts.append("updated_at=NOW()")
            vals.append(gid)

            cur.execute(
                f"UPDATE workmanship_tpl_gbop_entries SET {', '.join(set_parts)} "
                f"WHERE gid=%s",
                vals
            )
            if cur.rowcount == 0:
                raise HTTPException(404, "节点不存在")
            cur.execute(f"SELECT {_ENTRY_COLS} FROM workmanship_tpl_gbop_entries WHERE gid=%s", (gid,))
            row = cur.fetchone()
            conn.commit()
    return {"data": _entry_row(row)}


@router.delete("/entries/{gid}")
def delete_entry(gid: str, current_user: dict = Depends(_WRITE)):
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT version_gid FROM workmanship_tpl_gbop_entries WHERE gid=%s", (gid,))
            entry_row = cur.fetchone()
            if not entry_row:
                raise HTTPException(404, "节点不存在")
            _check_frozen(cur, entry_row['version_gid'])
            # 递归删除子节点（级联会自动删除 entry_links）
            cur.execute(
                "WITH RECURSIVE tree AS ("
                "  SELECT gid FROM workmanship_tpl_gbop_entries WHERE gid=%s "
                "  UNION ALL "
                "  SELECT e.gid FROM workmanship_tpl_gbop_entries e JOIN tree t ON e.parent_gid=t.gid"
                ") DELETE FROM workmanship_tpl_gbop_entries WHERE gid IN (SELECT gid FROM tree)",
                (gid,)
            )
            conn.commit()
    return {"success": True}


# ── Process CRUD ─────────────────────────────────────────────────

@router.get("/versions/{version_gid}/processes")
def list_processes(version_gid: str, current_user: dict = Depends(_READ)):
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute(
                f"SELECT {_PROC_COLS} FROM workmanship_tpl_gbop_processes "
                f"WHERE version_gid=%s ORDER BY created_at",
                (version_gid,)
            )
            rows = cur.fetchall()
    return {"data": [_entity_row(r) for r in rows]}


@router.post("/processes", status_code=201)
def create_process(body: CreateProcessBody, current_user: dict = Depends(_WRITE)):
    """一键创建：process + entry + link"""
    with get_conn() as conn:
        with conn.cursor() as cur:
            _check_frozen(cur, body.version_gid)

            proc_gid = str(next_gid())
            entry_gid = str(next_gid())
            link_gid = str(next_gid())
            level = _calc_level(cur, body.parent_entry_gid)

            # 1. 插入 gbop_processes
            cur.execute(
                "INSERT INTO workmanship_tpl_gbop_processes "
                "(gid, version_gid, vpps, vpps_desc, op_code, op_name, standard_time, "
                " description, steps, required_tools, parameters, "
                " importance, torque_importance, vehicle_model, meta, created_by, "
                " vpps_part, part_feed) "
                "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) ",
                (proc_gid, body.version_gid, body.vpps, body.vpps_desc,
                 body.op_code, body.op_name, body.standard_time,
                 body.description,
                 json.dumps(body.steps, ensure_ascii=False),
                 json.dumps(body.required_tools, ensure_ascii=False),
                 json.dumps(body.parameters, ensure_ascii=False),
                 body.importance, body.torque_importance, body.vehicle_model,
                 json.dumps(body.meta, ensure_ascii=False),
                 current_user["gid"],
                 body.vpps_part, body.part_feed)
            )
            cur.execute(f"SELECT {_PROC_COLS} FROM workmanship_tpl_gbop_processes WHERE gid=%s", (proc_gid,))
            process = _entity_row(cur.fetchone())

            # 2. 插入 gbop_entries
            cur.execute(
                "INSERT INTO workmanship_tpl_gbop_entries "
                "(gid, version_gid, parent_gid, level, node_type, seq_no, "
                " vpps, vpps_desc, importance, torque_importance, "
                " vehicle_model, meta, team_id, created_by, vpps_part, part_feed) "
                "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) ",
                (entry_gid, body.version_gid, body.parent_entry_gid, level, 'process', body.seq_no,
                 body.vpps, body.vpps_desc, body.importance, body.torque_importance,
                 body.vehicle_model, json.dumps(body.meta, ensure_ascii=False),
                 current_user.get("team_id"), current_user["gid"],
                 body.vpps_part, body.part_feed)
            )
            cur.execute(f"SELECT {_ENTRY_COLS} FROM workmanship_tpl_gbop_entries WHERE gid=%s", (entry_gid,))
            entry = _entry_row(cur.fetchone())

            # 3. 插入 gbop_entry_links
            cur.execute(
                f"INSERT INTO workmanship_tpl_gbop_entry_links "
                f"(gid, entry_gid, link_type, ref_gid, is_primary, created_by) "
                f"VALUES (%s,%s,%s,%s,%s,%s) ",
                (link_gid, entry_gid, 'gbop_process', proc_gid, True, current_user["gid"])
            )
            cur.execute(f"SELECT {_LINK_COLS} FROM workmanship_tpl_gbop_entry_links WHERE gid=%s", (link_gid,))
            link = _link_row(cur.fetchone())

            conn.commit()
    return {"data": {"process": process, "entry": entry, "link": link}}


@router.patch("/processes/{gid}")
def update_process(gid: str, body: UpdateProcessBody, current_user: dict = Depends(_WRITE)):
    data = body.model_dump(exclude_unset=True)
    if not data:
        raise HTTPException(400, "无更新字段")

    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT version_gid FROM workmanship_tpl_gbop_processes WHERE gid=%s", (gid,))
            proc = cur.fetchone()
            if not proc:
                raise HTTPException(404, "工艺卡片不存在")
            _check_frozen(cur, proc['version_gid'])

            json_fields = {'steps', 'required_tools', 'parameters', 'meta'}
            set_parts, vals = [], []
            for k, v in data.items():
                if k in json_fields:
                    set_parts.append(f"{k}=%s")
                    vals.append(json.dumps(v, ensure_ascii=False) if v is not None else '{}')
                else:
                    set_parts.append(f"{k}=%s")
                    vals.append(v)
            set_parts.append("updated_at=NOW()")
            vals.append(gid)

            cur.execute(
                f"UPDATE workmanship_tpl_gbop_processes SET {', '.join(set_parts)} "
                f"WHERE gid=%s",
                vals
            )
            cur.execute(f"SELECT {_PROC_COLS} FROM workmanship_tpl_gbop_processes WHERE gid=%s", (gid,))
            row = cur.fetchone()
            conn.commit()
    return {"data": _entity_row(row)}


@router.delete("/processes/{gid}")
def delete_process(gid: str, current_user: dict = Depends(_WRITE)):
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT version_gid FROM workmanship_tpl_gbop_processes WHERE gid=%s", (gid,))
            proc = cur.fetchone()
            if not proc:
                raise HTTPException(404, "工艺卡片不存在")
            _check_frozen(cur, proc['version_gid'])
            # 删除 process（entry_links 级联删除由 FK on ref_gid 不负责，需手动删）
            cur.execute(
                "DELETE FROM workmanship_tpl_gbop_entry_links WHERE link_type='gbop_process' AND ref_gid=%s",
                (gid,)
            )
            cur.execute("DELETE FROM workmanship_tpl_gbop_processes WHERE gid=%s", (gid,))
            conn.commit()
    return {"success": True}


# ── Operation CRUD ───────────────────────────────────────────────

@router.get("/versions/{version_gid}/operations")
def list_operations(version_gid: str, current_user: dict = Depends(_READ)):
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute(
                f"SELECT {_OP_COLS} FROM workmanship_tpl_gbop_operations "
                f"WHERE version_gid=%s ORDER BY created_at",
                (version_gid,)
            )
            rows = cur.fetchall()
    return {"data": [_entity_row(r) for r in rows]}


@router.post("/operations", status_code=201)
def create_operation(body: CreateOperationBody, current_user: dict = Depends(_WRITE)):
    """一键创建：operation + entry + link"""
    with get_conn() as conn:
        with conn.cursor() as cur:
            _check_frozen(cur, body.version_gid)

            op_gid = str(next_gid())
            entry_gid = str(next_gid())
            link_gid = str(next_gid())
            level = _calc_level(cur, body.parent_entry_gid)

            # 1. 插入 gbop_operations
            cur.execute(
                "INSERT INTO workmanship_tpl_gbop_operations "
                "(gid, version_gid, process_gid, vpps, vpps_desc, op_code, op_name, standard_time, "
                " description, steps, required_tools, parameters, "
                " importance, torque_importance, vehicle_model, meta, created_by, "
                " vpps_part, part_feed) "
                "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) ",
                (op_gid, body.version_gid, body.process_gid, body.vpps, body.vpps_desc,
                 body.op_code, body.op_name, body.standard_time,
                 body.description,
                 json.dumps(body.steps, ensure_ascii=False),
                 json.dumps(body.required_tools, ensure_ascii=False),
                 json.dumps(body.parameters, ensure_ascii=False),
                 body.importance, body.torque_importance, body.vehicle_model,
                 json.dumps(body.meta, ensure_ascii=False),
                 current_user["gid"],
                 body.vpps_part, body.part_feed)
            )
            cur.execute(f"SELECT {_OP_COLS} FROM workmanship_tpl_gbop_operations WHERE gid=%s", (op_gid,))
            operation = _entity_row(cur.fetchone())

            # 2. 插入 gbop_entries
            cur.execute(
                "INSERT INTO workmanship_tpl_gbop_entries "
                "(gid, version_gid, parent_gid, level, node_type, seq_no, "
                " vpps, vpps_desc, importance, torque_importance, "
                " vehicle_model, meta, team_id, created_by, vpps_part, part_feed) "
                "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) ",
                (entry_gid, body.version_gid, body.parent_entry_gid, level, 'operation', body.seq_no,
                 body.vpps, body.vpps_desc, body.importance, body.torque_importance,
                 body.vehicle_model, json.dumps(body.meta, ensure_ascii=False),
                 current_user.get("team_id"), current_user["gid"],
                 body.vpps_part, body.part_feed)
            )
            cur.execute(f"SELECT {_ENTRY_COLS} FROM workmanship_tpl_gbop_entries WHERE gid=%s", (entry_gid,))
            entry = _entry_row(cur.fetchone())

            # 3. 插入 gbop_entry_links
            cur.execute(
                f"INSERT INTO workmanship_tpl_gbop_entry_links "
                f"(gid, entry_gid, link_type, ref_gid, is_primary, created_by) "
                f"VALUES (%s,%s,%s,%s,%s,%s) ",
                (link_gid, entry_gid, 'gbop_operation', op_gid, True, current_user["gid"])
            )
            cur.execute(f"SELECT {_LINK_COLS} FROM workmanship_tpl_gbop_entry_links WHERE gid=%s", (link_gid,))
            link = _link_row(cur.fetchone())

            conn.commit()
    return {"data": {"operation": operation, "entry": entry, "link": link}}


@router.patch("/operations/{gid}")
def update_operation(gid: str, body: UpdateOperationBody, current_user: dict = Depends(_WRITE)):
    data = body.model_dump(exclude_unset=True)
    if not data:
        raise HTTPException(400, "无更新字段")

    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT version_gid FROM workmanship_tpl_gbop_operations WHERE gid=%s", (gid,))
            op = cur.fetchone()
            if not op:
                raise HTTPException(404, "操作卡片不存在")
            _check_frozen(cur, op['version_gid'])

            json_fields = {'steps', 'required_tools', 'parameters', 'meta'}
            set_parts, vals = [], []
            for k, v in data.items():
                if k in json_fields:
                    set_parts.append(f"{k}=%s")
                    vals.append(json.dumps(v, ensure_ascii=False) if v is not None else '{}')
                else:
                    set_parts.append(f"{k}=%s")
                    vals.append(v)
            set_parts.append("updated_at=NOW()")
            vals.append(gid)

            cur.execute(
                f"UPDATE workmanship_tpl_gbop_operations SET {', '.join(set_parts)} "
                f"WHERE gid=%s",
                vals
            )
            cur.execute(f"SELECT {_OP_COLS} FROM workmanship_tpl_gbop_operations WHERE gid=%s", (gid,))
            row = cur.fetchone()
            conn.commit()
    return {"data": _entity_row(row)}


@router.delete("/operations/{gid}")
def delete_operation(gid: str, current_user: dict = Depends(_WRITE)):
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT version_gid FROM workmanship_tpl_gbop_operations WHERE gid=%s", (gid,))
            op = cur.fetchone()
            if not op:
                raise HTTPException(404, "操作卡片不存在")
            _check_frozen(cur, op['version_gid'])
            cur.execute(
                "DELETE FROM workmanship_tpl_gbop_entry_links WHERE link_type='gbop_operation' AND ref_gid=%s",
                (gid,)
            )
            cur.execute("DELETE FROM workmanship_tpl_gbop_operations WHERE gid=%s", (gid,))
            conn.commit()
    return {"success": True}


# ── Entry Links CRUD ─────────────────────────────────────────────

@router.post("/entry-links", status_code=201)
def create_entry_link(body: CreateEntryLinkBody, current_user: dict = Depends(_WRITE)):
    gid = str(next_gid())
    with get_conn() as conn:
        with conn.cursor() as cur:
            # 检查 entry 存在且版本未冻结
            cur.execute("SELECT version_gid FROM workmanship_tpl_gbop_entries WHERE gid=%s", (body.entry_gid,))
            entry = cur.fetchone()
            if not entry:
                raise HTTPException(404, "节点不存在")
            _check_frozen(cur, entry['version_gid'])

            cur.execute(
                f"INSERT INTO workmanship_tpl_gbop_entry_links "
                f"(gid, entry_gid, link_type, ref_gid, is_primary, created_by) "
                f"VALUES (%s,%s,%s,%s,%s,%s) ",
                (gid, body.entry_gid, body.link_type, body.ref_gid,
                 body.is_primary, current_user["gid"])
            )
            cur.execute(f"SELECT {_LINK_COLS} FROM workmanship_tpl_gbop_entry_links WHERE gid=%s", (gid,))
            row = cur.fetchone()
            conn.commit()
    return {"data": _link_row(row)}


@router.delete("/entry-links/{gid}")
def delete_entry_link(gid: str, current_user: dict = Depends(_WRITE)):
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT l.entry_gid, e.version_gid "
                "FROM workmanship_tpl_gbop_entry_links l "
                "JOIN workmanship_tpl_gbop_entries e ON e.gid = l.entry_gid "
                "WHERE l.gid=%s",
                (gid,)
            )
            row = cur.fetchone()
            if not row:
                raise HTTPException(404, "链接不存在")
            _check_frozen(cur, row['version_gid'])
            cur.execute("DELETE FROM workmanship_tpl_gbop_entry_links WHERE gid=%s", (gid,))
            conn.commit()
    return {"success": True}


@router.get("/entries/{entry_gid}/links")
def get_entry_links(entry_gid: str, current_user: dict = Depends(_READ)):
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute(
                f"SELECT {_LINK_COLS} FROM workmanship_tpl_gbop_entry_links WHERE entry_gid=%s",
                (entry_gid,)
            )
            rows = cur.fetchall()
    return {"data": [_link_row(r) for r in rows]}


# ── Import ───────────────────────────────────────────────────────

@router.post("/versions/{version_gid}/import-vpps-parts", status_code=201)
def import_vpps_parts(version_gid: str, body: ImportVppsPartsBody, current_user: dict = Depends(_WRITE)):
    """从 workmanship_tpl_vpps_parts 导入 L1-3 节点，按 parent_vpps 匹配建立父子关系"""
    with get_conn() as conn:
        with conn.cursor() as cur:
            _check_frozen(cur, version_gid)

            # 查询 vpps_parts（active / 有效 均视为有效记录）
            cur.execute(
                "SELECT gid, vpps, vpps_description, vpps_desc_cn, level, importance, "
                "vehicle_model, part_category, parent_vpps, meta "
                "FROM workmanship_tpl_vpps_parts "
                "WHERE status IN ('active','有效') ORDER BY level, vpps"
            )
            parts = cur.fetchall()
            if not parts:
                cur.execute("SELECT count(*) AS cnt FROM workmanship_tpl_vpps_parts")
                total = cur.fetchone()['cnt']
                raise HTTPException(404, f"vpps_parts 中无有效记录（共 {total} 条）")

            # 按 vpps 建立索引，用于查找父节点
            vpps_to_entry_gid = {}
            created = 0
            level_map = {'1': 1, '2': 2, '3': 3}

            for part in parts:
                part_level = level_map.get(str(part['level']).strip())
                if part_level is None or part_level not in body.levels:
                    continue

                node_type_map = {1: 'system', 2: 'device', 3: 'part'}
                node_type = node_type_map.get(part_level, 'part')

                # 查找父级：优先用独立列 parent_vpps，fallback 到 meta
                parent_vpps_val = (part.get('parent_vpps') or '').strip()
                if not parent_vpps_val:
                    raw_meta = part['meta']
                    if isinstance(raw_meta, str):
                        try:
                            raw_meta = json.loads(raw_meta)
                        except Exception:
                            raw_meta = {}
                    part_meta = raw_meta if isinstance(raw_meta, dict) else {}
                    parent_vpps_val = part_meta.get('parent_vpps', '')
                parent_entry_gid = vpps_to_entry_gid.get(parent_vpps_val) if parent_vpps_val else None

                entry_gid = str(next_gid())
                vpps_val = part['vpps'] or ''
                vpps_desc = part.get('vpps_description') or part.get('vpps_desc_cn') or ''
                vpps_attr = part.get('part_category') or ''

                cur.execute(
                    "INSERT INTO workmanship_tpl_gbop_entries "
                    "(gid, version_gid, parent_gid, level, node_type, seq_no, "
                    " vpps, vpps_desc, vpps_attr, importance, vehicle_model, "
                    " parent_vpps, team_id, created_by) "
                    "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                    (entry_gid, version_gid, parent_entry_gid, part_level,
                     node_type, created,
                     vpps_val, vpps_desc, vpps_attr,
                     part.get('importance') or '',
                     part.get('vehicle_model') or '',
                     parent_vpps_val,
                     current_user.get("team_id"), current_user["gid"])
                )
                vpps_to_entry_gid[vpps_val] = entry_gid
                created += 1

            conn.commit()
    return {"data": {"created_count": created}}


@router.post("/versions/{version_gid}/import-entries", status_code=201)
def import_entries(version_gid: str, body: ImportEntriesBody, current_user: dict = Depends(_WRITE)):
    """批量导入条目（从前端解析后的 JSON 数组）"""
    with get_conn() as conn:
        with conn.cursor() as cur:
            _check_frozen(cur, version_gid)

            # 第一遍：分配 gid，建立 parent_vpps → gid 映射
            gid_map = {}  # vpps → new_gid
            entries_with_gid = []
            for entry in body.entries:
                new_gid = str(next_gid())
                vpps_val = entry.get('vpps') or ''
                if vpps_val:
                    gid_map[vpps_val] = new_gid
                entries_with_gid.append((new_gid, entry))

            # 第二遍：插入，parent_vpps 匹配建立关系
            created = 0
            for new_gid, entry in entries_with_gid:
                parent_vpps_val = entry.get('parent_vpps', '')
                parent_gid = entry.get('parent_gid') or gid_map.get(parent_vpps_val)
                level = int(entry.get('level', 0))
                node_type = entry.get('node_type', 'process')

                cur.execute(
                    "INSERT INTO workmanship_tpl_gbop_entries "
                    "(gid, version_gid, parent_gid, level, node_type, seq_no, "
                    " vpps, vpps_desc, vpps_attr, importance, torque_importance, "
                    " vehicle_model, parent_vpps, meta, "
                    " team_id, created_by) "
                    "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,"
                    "        %s,%s,%s)",
                    (new_gid, version_gid, parent_gid, level, node_type,
                     entry.get('seq_no', created),
                     entry.get('vpps'), entry.get('vpps_desc', ''),
                     entry.get('vpps_attr', ''), entry.get('importance', ''),
                     entry.get('torque_importance', ''), entry.get('vehicle_model', ''),
                     parent_vpps_val,
                     json.dumps(entry.get('meta', {}), ensure_ascii=False),
                     current_user.get("team_id"), current_user["gid"])
                )
                created += 1

            conn.commit()
    return {"data": {"created_count": created}}


# ── TC Excel Import ───────────────────────────────────────────────

@router.post("/versions/{version_gid}/import-tc-excel", status_code=201)
async def import_tc_excel(
    version_gid: str,
    file: UploadFile = File(...),
    current_user: dict = Depends(_WRITE),
):
    """从 Teamcenter 导出的两 Sheet Excel 导入工序/操作，并绑定到零件 VPPS。

    Sheet 1 — 工序/操作清单：
      只处理 零组件类型="总装工序"（→ gbop_processes）和
              零组件类型 含 "总装操作"（→ gbop_operations）的行。
      通过 BOM行/父级 列建立操作→工序父子关系，自动填写 process_gid。

    Sheet 2 — VPPS 绑定表：
      将 gbop_entry(part, vpps) → gbop_process/gbop_operation 的链接
      写入 gbop_entry_links。
    """
    import openpyxl

    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise HTTPException(400, f"无法解析 Excel 文件：{exc}")

    if '1' not in wb.sheetnames:
        raise HTTPException(400, "Excel 缺少工作表 '1'")

    ws1 = wb['1']
    rows1 = list(ws1.iter_rows(values_only=True))
    if len(rows1) < 2:
        raise HTTPException(400, "Sheet 1 数据为空")

    # ── 列索引 ──
    h1 = [str(c).strip() if c is not None else '' for c in rows1[0]]

    def _ci(headers, *names):
        for n in names:
            try:
                return headers.index(n)
            except ValueError:
                pass
        return -1

    c1_type   = _ci(h1, '零组件类型')
    c1_name   = _ci(h1, '零组件名称')
    c1_bom    = _ci(h1, 'BOM 行')
    c1_vpps   = _ci(h1, 'VPPS')
    c1_parent = _ci(h1, '父级')

    def _val(row, idx):
        return str(row[idx]).strip() if idx >= 0 and row[idx] is not None else ''

    # ── 分拣 Sheet 1 行 ──
    proc_rows = []
    op_rows   = []
    for row in rows1[1:]:
        rt = _val(row, c1_type)
        if rt == '总装工序':
            proc_rows.append(row)
        elif '总装操作' in rt:
            op_rows.append(row)

    with get_conn() as conn:
        with conn.cursor() as cur:
            _check_frozen(cur, version_gid)

            # ── Phase 1a：写入 gbop_processes ──
            bom_to_proc: Dict[str, dict] = {}   # bom_行字符串 → {gid, vpps, name}
            proc_vpps_to_gid: Dict[str, str] = {}
            proc_vpps_to_name: Dict[str, str] = {}
            proc_created = 0

            for row in proc_rows:
                vpps = _val(row, c1_vpps)
                name = _val(row, c1_name)
                bom  = _val(row, c1_bom)
                if not vpps:
                    continue
                gid = str(next_gid())
                cur.execute(
                    "INSERT INTO workmanship_tpl_gbop_processes "
                    "(gid, version_gid, vpps, vpps_desc, op_code, op_name, created_by) "
                    "VALUES (%s,%s,%s,%s,%s,%s,%s)",
                    (gid, version_gid, vpps, name, vpps, name, current_user['gid']),
                )
                proc_vpps_to_gid[vpps] = gid
                proc_vpps_to_name[vpps] = name
                if bom:
                    bom_to_proc[bom] = {'gid': gid, 'vpps': vpps, 'name': name}
                proc_created += 1

            # ── Phase 1b：写入 gbop_operations ──
            op_vpps_to_gid:  Dict[str, str] = {}
            op_vpps_to_name: Dict[str, str] = {}
            op_created = 0

            for row in op_rows:
                vpps        = _val(row, c1_vpps)
                name        = _val(row, c1_name)
                parent_bom  = _val(row, c1_parent)
                if not vpps:
                    continue
                parent_proc = bom_to_proc.get(parent_bom)
                process_gid = parent_proc['gid'] if parent_proc else None
                gid = str(next_gid())
                cur.execute(
                    "INSERT INTO workmanship_tpl_gbop_operations "
                    "(gid, version_gid, process_gid, vpps, vpps_desc, op_code, op_name, created_by) "
                    "VALUES (%s,%s,%s,%s,%s,%s,%s,%s)",
                    (gid, version_gid, process_gid, vpps, name, vpps, name, current_user['gid']),
                )
                op_vpps_to_gid[vpps]  = gid
                op_vpps_to_name[vpps] = name
                op_created += 1

            # ── Phase 2：Sheet 2 → gbop_entries（process/operation）+ gbop_entry_links ──
            links_created   = 0
            entries_created = 0
            # (part_entry_gid, proc_vpps) → proc_entry_gid（每对唯一，避免重复建节点）
            proc_entry_map: Dict[tuple, str] = {}
            # (proc_entry_gid, op_vpps)  → op_entry_gid
            op_entry_map:   Dict[tuple, str] = {}
            # entry_gid → [{vpps, node_type, title}] 用于回写 child_vpps
            entry_child_map: Dict[str, list] = {}

            if '2' in wb.sheetnames:
                ws2   = wb['2']
                rows2 = list(ws2.iter_rows(values_only=True))
                if len(rows2) >= 2:
                    h2      = [str(c).strip() if c is not None else '' for c in rows2[0]]
                    c2_vpps = _ci(h2, 'VPPS')
                    c2_proc = _ci(h2, '工序VPPS')
                    c2_op   = _ci(h2, '操作VPPS')
                    c2_tag  = _ci(h2, '标记')

                    # seq 计数器（按父节点分别累加）
                    proc_seq: Dict[str, int] = {}
                    op_seq:   Dict[str, int] = {}

                    def _lookup_proc_entity(vpps_val):
                        g = proc_vpps_to_gid.get(vpps_val)
                        if g:
                            return g
                        cur.execute(
                            "SELECT gid FROM workmanship_tpl_gbop_processes "
                            "WHERE version_gid=%s AND vpps=%s LIMIT 1",
                            (version_gid, vpps_val),
                        )
                        r = cur.fetchone()
                        return r['gid'] if r else None

                    def _lookup_op_entity(vpps_val):
                        g = op_vpps_to_gid.get(vpps_val)
                        if g:
                            return g
                        cur.execute(
                            "SELECT gid FROM workmanship_tpl_gbop_operations "
                            "WHERE version_gid=%s AND vpps=%s LIMIT 1",
                            (version_gid, vpps_val),
                        )
                        r = cur.fetchone()
                        return r['gid'] if r else None

                    for row in rows2[1:]:
                        part_vpps    = _val(row, c2_vpps)
                        proc_vpps    = _val(row, c2_proc)
                        op_vpps      = _val(row, c2_op)
                        is_part_feed = _val(row, c2_tag).lower() == 'part_feed'
                        if not proc_vpps and not op_vpps:
                            continue

                        # ── 找 part gbop_entry（找不到则以顶层挂载）──
                        part_entry_gid = None
                        part_level     = -1   # process=0, operation=1
                        if part_vpps:
                            cur.execute(
                                "SELECT gid, level FROM workmanship_tpl_gbop_entries "
                                "WHERE version_gid=%s AND vpps=%s AND node_type='part' LIMIT 1",
                                (version_gid, part_vpps),
                            )
                            part_r = cur.fetchone()
                            if part_r:
                                part_entry_gid = part_r['gid']
                                part_level     = part_r['level']

                        # ── 创建/复用 process entry（每个 part×proc 组合唯一）──
                        proc_entry_gid = None
                        if proc_vpps:
                            proc_key = (part_entry_gid, proc_vpps)
                            if proc_key in proc_entry_map:
                                proc_entry_gid = proc_entry_map[proc_key]
                            else:
                                p_entity_gid = _lookup_proc_entity(proc_vpps)
                                if p_entity_gid:
                                    seq = proc_seq.get(part_entry_gid, 0)
                                    proc_seq[part_entry_gid] = seq + 1
                                    proc_entry_gid = str(next_gid())
                                    proc_name = proc_vpps_to_name.get(proc_vpps, proc_vpps)
                                    cur.execute(
                                        "INSERT INTO workmanship_tpl_gbop_entries "
                                        "(gid, version_gid, parent_gid, level, node_type, seq_no, "
                                        " vpps, vpps_desc, meta, team_id, created_by, vpps_part, part_feed) "
                                        "VALUES (%s,%s,%s,%s,'process',%s,%s,%s,'{}',%s,%s,%s,FALSE)",
                                        (proc_entry_gid, version_gid, part_entry_gid,
                                         part_level + 1, seq,
                                         proc_vpps, proc_name,
                                         current_user.get('team_id'), current_user['gid'],
                                         part_vpps),
                                    )
                                    cur.execute(
                                        "INSERT INTO workmanship_tpl_gbop_entry_links "
                                        "(gid, entry_gid, link_type, ref_gid, is_primary, created_by) "
                                        "VALUES (%s,%s,'gbop_process',%s,TRUE,%s)",
                                        (str(next_gid()), proc_entry_gid, p_entity_gid, current_user['gid']),
                                    )
                                    proc_entry_map[proc_key] = proc_entry_gid
                                    entries_created += 1
                                    links_created   += 1
                                    cv = entry_child_map.setdefault(part_entry_gid, [])
                                    if not any(c['vpps'] == proc_vpps for c in cv):
                                        cv.append({'vpps': proc_vpps, 'node_type': 'process', 'title': proc_name})

                        # ── 创建/复用 operation entry（每个 proc_entry×op 组合唯一）──
                        if op_vpps and proc_entry_gid:
                            op_key = (proc_entry_gid, op_vpps)
                            if op_key not in op_entry_map:
                                o_entity_gid = _lookup_op_entity(op_vpps)
                                if o_entity_gid:
                                    seq = op_seq.get(proc_entry_gid, 0)
                                    op_seq[proc_entry_gid] = seq + 1
                                    op_entry_gid = str(next_gid())
                                    op_name = op_vpps_to_name.get(op_vpps, op_vpps)
                                    cur.execute(
                                        "INSERT INTO workmanship_tpl_gbop_entries "
                                        "(gid, version_gid, parent_gid, level, node_type, seq_no, "
                                        " vpps, vpps_desc, meta, team_id, created_by, vpps_part, part_feed) "
                                        "VALUES (%s,%s,%s,%s,'operation',%s,%s,%s,'{}',%s,%s,%s,%s)",
                                        (op_entry_gid, version_gid, proc_entry_gid,
                                         part_level + 2, seq,
                                         op_vpps, op_name,
                                         current_user.get('team_id'), current_user['gid'],
                                         part_vpps, is_part_feed),
                                    )
                                    cur.execute(
                                        "INSERT INTO workmanship_tpl_gbop_entry_links "
                                        "(gid, entry_gid, link_type, ref_gid, is_primary, created_by) "
                                        "VALUES (%s,%s,'gbop_operation',%s,TRUE,%s)",
                                        (str(next_gid()), op_entry_gid, o_entity_gid, current_user['gid']),
                                    )
                                    op_entry_map[op_key] = op_entry_gid
                                    entries_created += 1
                                    links_created   += 1
                                    if is_part_feed:
                                        cur.execute(
                                            "UPDATE workmanship_tpl_gbop_operations SET part_feed=TRUE WHERE gid=%s",
                                            (o_entity_gid,),
                                        )

            # ── 回写 child_vpps 到零件 entry ──
            for e_gid, child_list in entry_child_map.items():
                cur.execute(
                    "UPDATE workmanship_tpl_gbop_entries SET child_vpps=%s, updated_at=NOW() WHERE gid=%s",
                    (json.dumps(child_list), e_gid),
                )

    return {"data": {
        "processes_created":  proc_created,
        "operations_created": op_created,
        "entries_created":    entries_created,
        "links_created":      links_created,
    }}


# ── Fork ─────────────────────────────────────────────────────────

@router.post("/versions/{source_gid}/fork", status_code=201)
def fork_version(source_gid: str, body: ForkBody, current_user: dict = Depends(_WRITE)):
    """Fork 版本：复制 entries + processes + operations + entry_links，重映射所有 gid"""
    with get_conn() as conn:
        with conn.cursor() as cur:
            # 确认源版本存在
            cur.execute(f"SELECT {_VER_COLS} FROM workmanship_tpl_gbop_versions WHERE gid=%s", (source_gid,))
            src_ver = cur.fetchone()
            if not src_ver:
                raise HTTPException(404, "源版本不存在")

            # 创建新版本
            new_ver_gid = str(next_gid())
            family_gid = body.target_version_family_gid or src_ver['version_family_gid']
            name = body.target_name or src_ver['name']

            cur.execute(
                f"INSERT INTO workmanship_tpl_gbop_versions "
                f"(gid, name, version_family_gid, vehicle_model, team_id, created_by) "
                f"VALUES (%s,%s,%s,%s,%s,%s)",
                (new_ver_gid, name, family_gid,
                 src_ver.get('vehicle_model') or '',
                 current_user.get("team_id"), current_user["gid"])
            )
            cur.execute(f"SELECT {_VER_COLS} FROM workmanship_tpl_gbop_versions WHERE gid=%s", (new_ver_gid,))
            new_ver = _ver_row(cur.fetchone())

            # ── 1. 复制 entries ──
            cur.execute(
                f"SELECT {_ENTRY_COLS} FROM workmanship_tpl_gbop_entries "
                f"WHERE version_gid=%s ORDER BY level, seq_no",
                (source_gid,)
            )
            src_entries = cur.fetchall()

            entry_remap = {}  # old_gid → new_gid
            for entry in src_entries:
                if body.include_node_types and entry['node_type'] not in body.include_node_types:
                    continue
                entry_remap[entry['gid']] = str(next_gid())

            for entry in src_entries:
                old_gid = entry['gid']
                if old_gid not in entry_remap:
                    continue
                new_entry_gid = entry_remap[old_gid]
                new_parent = entry_remap.get(entry['parent_gid']) if entry['parent_gid'] else None

                cur.execute(
                    "INSERT INTO workmanship_tpl_gbop_entries "
                    "(gid, version_gid, parent_gid, level, node_type, seq_no, "
                    " vpps, vpps_desc, vpps_attr, importance, torque_importance, "
                    " vehicle_model, parent_vpps, status, sort_order, "
                    " meta, team_id, created_by, vpps_part, part_feed) "
                    "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,"
                    "        %s,%s,%s,%s,%s)",
                    (new_entry_gid, new_ver_gid, new_parent,
                     entry['level'], entry['node_type'], entry['seq_no'],
                     entry['vpps'], entry['vpps_desc'], entry['vpps_attr'],
                     entry['importance'], entry['torque_importance'],
                     entry['vehicle_model'], entry['parent_vpps'],
                     entry['status'], entry['sort_order'],
                     json.dumps(entry['meta'] or {}, ensure_ascii=False),
                     current_user.get("team_id"), current_user["gid"],
                     entry.get('vpps_part', ''), entry.get('part_feed', False))
                )

            # ── 2. 复制 processes ──
            cur.execute(
                f"SELECT {_PROC_COLS} FROM workmanship_tpl_gbop_processes WHERE version_gid=%s",
                (source_gid,)
            )
            src_procs = cur.fetchall()

            proc_remap = {}  # old_gid → new_gid
            for p in src_procs:
                proc_remap[p['gid']] = str(next_gid())

            for p in src_procs:
                new_proc_gid = proc_remap[p['gid']]
                cur.execute(
                    "INSERT INTO workmanship_tpl_gbop_processes "
                    "(gid, version_gid, vpps, vpps_desc, op_code, op_name, standard_time, "
                    " description, steps, required_tools, parameters, "
                    " importance, torque_importance, vehicle_model, status, meta, created_by, "
                    " vpps_part, part_feed) "
                    "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                    (new_proc_gid, new_ver_gid, p['vpps'], p['vpps_desc'],
                     p['op_code'], p['op_name'], p['standard_time'],
                     p['description'],
                     json.dumps(p['steps'] or [], ensure_ascii=False),
                     json.dumps(p['required_tools'] or [], ensure_ascii=False),
                     json.dumps(p['parameters'] or {}, ensure_ascii=False),
                     p['importance'], p['torque_importance'], p['vehicle_model'],
                     p['status'],
                     json.dumps(p['meta'] or {}, ensure_ascii=False),
                     current_user["gid"],
                     p.get('vpps_part', ''), p.get('part_feed', False))
                )

            # ── 3. 复制 operations ──
            cur.execute(
                f"SELECT {_OP_COLS} FROM workmanship_tpl_gbop_operations WHERE version_gid=%s",
                (source_gid,)
            )
            src_ops = cur.fetchall()

            op_remap = {}  # old_gid → new_gid
            for o in src_ops:
                op_remap[o['gid']] = str(next_gid())

            for o in src_ops:
                new_op_gid = op_remap[o['gid']]
                new_process_gid = proc_remap.get(o['process_gid']) if o['process_gid'] else None
                cur.execute(
                    "INSERT INTO workmanship_tpl_gbop_operations "
                    "(gid, version_gid, process_gid, vpps, vpps_desc, op_code, op_name, standard_time, "
                    " description, steps, required_tools, parameters, "
                    " importance, torque_importance, vehicle_model, status, meta, created_by, "
                    " vpps_part, part_feed) "
                    "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                    (new_op_gid, new_ver_gid, new_process_gid, o['vpps'], o['vpps_desc'],
                     o['op_code'], o['op_name'], o['standard_time'],
                     o['description'],
                     json.dumps(o['steps'] or [], ensure_ascii=False),
                     json.dumps(o['required_tools'] or [], ensure_ascii=False),
                     json.dumps(o['parameters'] or {}, ensure_ascii=False),
                     o['importance'], o['torque_importance'], o['vehicle_model'],
                     o['status'],
                     json.dumps(o['meta'] or {}, ensure_ascii=False),
                     current_user["gid"],
                     o.get('vpps_part', ''), o.get('part_feed', False))
                )

            # ── 4. 复制 entry_links ──
            cur.execute(
                f"SELECT {_LINK_COLS} FROM workmanship_tpl_gbop_entry_links "
                f"WHERE entry_gid IN (SELECT gid FROM workmanship_tpl_gbop_entries WHERE version_gid=%s)",
                (source_gid,)
            )
            src_links = cur.fetchall()

            for lk in src_links:
                old_entry = lk['entry_gid']
                new_entry = entry_remap.get(old_entry)
                if not new_entry:
                    continue  # entry 被过滤掉了

                # 重映射 ref_gid
                lt = lk['link_type']
                old_ref = lk['ref_gid']
                if lt == 'gbop_process':
                    new_ref = proc_remap.get(old_ref)
                elif lt == 'gbop_operation':
                    new_ref = op_remap.get(old_ref)
                else:
                    new_ref = old_ref  # 未知类型保持原样

                if not new_ref:
                    continue

                new_link_gid = str(next_gid())
                cur.execute(
                    f"INSERT INTO workmanship_tpl_gbop_entry_links "
                    f"(gid, entry_gid, link_type, ref_gid, is_primary, created_by) "
                    f"VALUES (%s,%s,%s,%s,%s,%s)",
                    (new_link_gid, new_entry, lt, new_ref,
                     lk['is_primary'], current_user["gid"])
                )

            conn.commit()

    total_copied = len(entry_remap) + len(proc_remap) + len(op_remap) + len(src_links)
    return {"data": {
        "version": new_ver,
        "entries_copied": len(entry_remap),
        "processes_copied": len(proc_remap),
        "operations_copied": len(op_remap),
        "links_copied": len(src_links),
    }}


# ══════════════════════════════════════════════════════════════════
# GBOP 车型 Auto-Link：PBOM vpps → workmanship_tpl_gbop_entries 匹配
# ══════════════════════════════════════════════════════════════════

@router.get("/pbom-versions/{pbom_gid}/gbop-nav-link-summary")
def gbop_nav_link_summary(pbom_gid: str, _u=Depends(_READ)):
    """
    返回 gbop_nav_bindings 的 link-summary 格式：
    { gbop_op_entry_gid: { bop_entry_gid, is_valid, pbom_entry_gid } }
    供 AssocPanel 的 gbop_nav 适配器使用。
    """
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT gbop_op_entry_gid, pbom_entry_gid, confirmed
                FROM workmanship_bop_gbop_nav_bindings
                WHERE pbom_version_gid = %s
                """,
                (pbom_gid,),
            )
            rows = cur.fetchall()
    link_map: dict = {}
    for r in rows:
        op_gid = r['gbop_op_entry_gid']
        if op_gid not in link_map:
            link_map[op_gid] = {
                'bop_entry_gid': r['pbom_entry_gid'],
                'is_valid': True,
            }
    return {"data": link_map}


@router.get("/pbom-versions/{pbom_gid}/vpps-auto-link-status")
def gbop_vpps_auto_link_status(pbom_gid: str, _u=Depends(_READ)):
    """
    返回当前 pbom 版本的 Auto-Link 状态：
    pending_count > 0 表示有未提交绑定，前端应禁止再次 Auto-Link。
    """
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT COUNT(*) AS cnt FROM workmanship_bop_gbop_nav_bindings "
                "WHERE pbom_version_gid=%s AND confirmed=FALSE",
                (pbom_gid,),
            )
            pending_count = cur.fetchone()['cnt']
            cur.execute(
                "SELECT COUNT(*) AS cnt FROM workmanship_bop_gbop_nav_bindings "
                "WHERE pbom_version_gid=%s AND confirmed=TRUE",
                (pbom_gid,),
            )
            confirmed_count = cur.fetchone()['cnt']
    return {"data": {"pending_count": pending_count, "confirmed_count": confirmed_count}}


@router.post("/pbom-versions/{pbom_gid}/vpps-auto-link-confirm", status_code=200)
def gbop_vpps_auto_link_confirm(pbom_gid: str, _u=Depends(_WRITE)):
    """
    将当前 pbom 版本所有未提交(confirmed=FALSE)的 Auto-Link 绑定整体确认。
    确认后 Auto-Link 可再次执行。
    """
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "UPDATE workmanship_bop_gbop_nav_bindings SET confirmed=TRUE "
                "WHERE pbom_version_gid=%s AND confirmed=FALSE",
                (pbom_gid,),
            )
            updated = cur.rowcount
        conn.commit()
    return {"ok": True, "confirmed": updated}


@router.post("/pbom-versions/{pbom_gid}/vpps-auto-link", status_code=200)
def gbop_vpps_auto_link(pbom_gid: str, current_user: dict = Depends(_WRITE)):
    """
    Auto-Link：遍历 PBOM 零件的 vpps，
    在 workmanship_tpl_gbop_entries 中找 vpps_part 相同且 part_feed=TRUE 的操作节点，
    向上溯源找到父工序节点，批量写入 workmanship_bop_gbop_nav_bindings（幂等）。
    """
    # 有未提交绑定时拒绝重复执行
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT COUNT(*) AS cnt FROM workmanship_bop_gbop_nav_bindings "
                "WHERE pbom_version_gid=%s AND confirmed=FALSE",
                (pbom_gid,),
            )
            pending = cur.fetchone()['cnt']
    if pending > 0:
        raise HTTPException(
            status_code=409,
            detail=f"存在 {pending} 条未提交的 Auto-Link 绑定，请先点击「确认绑定」后再执行。",
        )
    with get_conn() as conn:
        with conn.cursor() as cur:
            # 1. 取 PBOM 零件
            cur.execute(
                "SELECT gid, vpps, title, part_no "
                "FROM workmanship_bop_pbom "
                "WHERE snapshot_gid=%s AND COALESCE(is_deleted,FALSE)=FALSE AND vpps IS NOT NULL AND vpps != ''",
                (pbom_gid,),
            )
            parts = [dict(r) for r in cur.fetchall()]
            if not parts:
                return {"data": {"bound": 0, "parts_matched": 0}}

            part_vpps_list = list({p['vpps'] for p in parts})

            # 2. 取所有活跃 GBOP operation entries 中 vpps_part 匹配的
            if not part_vpps_list:
                op_entries = []
            else:
                _ph = ",".join(["%s"] * len(part_vpps_list))
                cur.execute(
                    f"""
                SELECT e.gid AS entry_gid, e.vpps, e.vpps_desc, e.part_feed,
                       e.parent_gid, e.node_type, e.version_gid
                FROM workmanship_tpl_gbop_entries e
                JOIN workmanship_tpl_gbop_versions v ON v.gid = e.version_gid
                WHERE e.node_type = 'operation'
                  AND e.part_feed = TRUE
                  AND e.vpps_part IN ({_ph})
                  AND v.archived_at IS NULL
                """,
                    part_vpps_list,
                )
                op_entries = [dict(r) for r in cur.fetchall()]

            if not op_entries:
                return {"data": {"bound": 0, "parts_matched": 0}}

            # 3. 批量取这些 operation 的父工序（node_type='process'）
            op_parent_gids = list({e['parent_gid'] for e in op_entries if e['parent_gid']})
            proc_map: dict = {}  # entry_gid → entry row
            if op_parent_gids:
                _ph = ",".join(["%s"] * len(op_parent_gids))
                cur.execute(
                    f"SELECT gid, vpps, vpps_desc, node_type, parent_gid "
                    f"FROM workmanship_tpl_gbop_entries "
                    f"WHERE gid IN ({_ph}) AND node_type = 'process'",
                    op_parent_gids,
                )
                for r in cur.fetchall():
                    proc_map[r['gid']] = dict(r)

            # 4. 建 vpps_part → op_entries 映射
            vpps_to_ops: dict = {}
            for e in op_entries:
                # 找 vpps_part 字段：需要重新查（上面 SELECT 没取 vpps_part）
                pass

            # 重新取含 vpps_part 的操作 entry
            if not part_vpps_list:
                op_entries = []
            else:
                _ph = ",".join(["%s"] * len(part_vpps_list))
                cur.execute(
                    f"""
                SELECT e.gid AS entry_gid, e.vpps, e.vpps_desc, e.vpps_part,
                       e.part_feed, e.parent_gid, e.node_type
                FROM workmanship_tpl_gbop_entries e
                JOIN workmanship_tpl_gbop_versions v ON v.gid = e.version_gid
                WHERE e.node_type = 'operation'
                  AND e.part_feed = TRUE
                  AND e.vpps_part IN ({_ph})
                  AND v.archived_at IS NULL
                """,
                    part_vpps_list,
                )
                op_entries = [dict(r) for r in cur.fetchall()]

            vpps_to_ops = {}
            for e in op_entries:
                vpps_to_ops.setdefault(e['vpps_part'], []).append(e)

            # 5. 逐零件写入 gbop_nav_bindings
            bound = 0
            parts_matched = 0
            for part in parts:
                vpps = part['vpps']
                matched_ops = vpps_to_ops.get(vpps, [])
                if not matched_ops:
                    continue
                parts_matched += 1
                for op in matched_ops:
                    proc_gid = op.get('parent_gid')
                    # 确保 proc_gid 确实是 process 节点
                    if proc_gid and proc_gid not in proc_map:
                        proc_gid = None
                    cur.execute(
                        """
                        INSERT INTO workmanship_bop_gbop_nav_bindings
                            (gid, pbom_version_gid, gbop_process_entry_gid,
                             gbop_op_entry_gid, pbom_entry_gid, is_part_feed)
                        VALUES (%s,%s,%s,%s,%s,TRUE)
                        ON DUPLICATE KEY UPDATE
                            gbop_process_entry_gid = VALUES(gbop_process_entry_gid),
                            is_part_feed = TRUE
                        """,
                        (str(next_gid()), pbom_gid,
                         proc_gid, op['entry_gid'], part['gid']),
                    )
                    bound += 1

        conn.commit()

    return {"data": {"bound": bound, "parts_matched": parts_matched}}


@router.get("/pbom-versions/{pbom_gid}/process-hierarchy")
def gbop_process_hierarchy(pbom_gid: str, _u=Depends(_READ)):
    """
    工序视图：返回三级树 process → operation → [parts]。
    骨架来自 workmanship_tpl_gbop_entries（全部 process/operation），
    零件绑定来自 workmanship_bop_gbop_nav_bindings（仅 part_feed=TRUE 的操作才有零件）。
    """
    with get_conn() as conn:
        with conn.cursor() as cur:
            # 1. 取所有活跃 GBOP 版本的 process / operation entry
            cur.execute(
                """
                SELECT e.gid, e.vpps, e.vpps_desc, e.node_type,
                       e.seq_no, e.parent_gid, e.part_feed
                FROM workmanship_tpl_gbop_entries e
                JOIN workmanship_tpl_gbop_versions v ON v.gid = e.version_gid
                WHERE e.node_type IN ('process', 'operation')
                  AND v.archived_at IS NULL
                ORDER BY e.seq_no 
                """,
            )
            all_entries = [dict(r) for r in cur.fetchall()]

            if not all_entries:
                return {"data": []}

            # 2. 取该 PBOM 版本所有绑定
            cur.execute(
                """
                SELECT gbop_op_entry_gid, pbom_entry_gid, confirmed
                FROM workmanship_bop_gbop_nav_bindings
                WHERE pbom_version_gid = %s
                """,
                (pbom_gid,),
            )
            bindings = [dict(r) for r in cur.fetchall()]

            # 3. 查 workmanship_bop_pbom 零件信息
            part_gids = list({b['pbom_entry_gid'] for b in bindings})
            part_map: dict = {}
            if part_gids:
                _ph = ",".join(["%s"] * len(part_gids))
                cur.execute(
                    f"SELECT gid, vpps, title, part_no FROM workmanship_bop_pbom WHERE gid IN ({_ph})",
                    part_gids,
                )
                for r in cur.fetchall():
                    part_map[r['gid']] = dict(r)

            # 4. 建 op_gid → parts 映射
            op_parts: dict = {}
            for b in bindings:
                og = b['gbop_op_entry_gid']
                part_row = part_map.get(b['pbom_entry_gid'])
                if part_row:
                    op_parts.setdefault(og, []).append({
                        'pbom_entry_gid': b['pbom_entry_gid'],
                        'vpps':      part_row.get('vpps', ''),
                        'title':     part_row.get('title', ''),
                        'part_no':   part_row.get('part_no', ''),
                        'confirmed': b['confirmed'],
                    })

            # 5. 建 entry_map 与 proc→ops 树
            entry_map = {e['gid']: e for e in all_entries}

            proc_ops: dict = {}   # proc_gid → [op_entry]

            for e in all_entries:
                if e['node_type'] != 'operation':
                    continue
                pg = e.get('parent_gid')
                if pg and pg in entry_map and entry_map[pg]['node_type'] == 'process':
                    proc_ops.setdefault(pg, []).append(e)

            # 6. 组装结果
            result = []
            for e in all_entries:
                if e['node_type'] != 'process':
                    continue
                pg = e['gid']
                operations = []
                for op in proc_ops.get(pg, []):
                    operations.append({
                        'entry_gid': op['gid'],
                        'vpps':      op.get('vpps', ''),
                        'title':     op.get('vpps_desc') or op.get('vpps', op['gid']),
                        'seq_no':    op.get('seq_no') or 0,
                        'part_feed': op.get('part_feed', False),
                        'parts':     op_parts.get(op['gid'], []),
                    })
                operations.sort(key=lambda x: x['seq_no'])
                result.append({
                    'process_entry_gid': pg,
                    'vpps':       e.get('vpps', ''),
                    'title':      e.get('vpps_desc') or e.get('vpps', '（无工序）'),
                    'seq_no':     e.get('seq_no') or 0,
                    'operations': operations,
                    'op_count':   len(operations),
                    'part_count': sum(len(o['parts']) for o in operations),
                })

            result.sort(key=lambda x: x['seq_no'])
            return {"data": result}


# ══════════════════════════════════════════════════════════════════
# BOP 版本 → GBOP nav bindings 工位自动关联（预览 + 执行）
# ══════════════════════════════════════════════════════════════════

@router.get("/bop-versions/{bop_gid}/station-autolink-preview")
def station_autolink_preview(
    bop_gid: str,
    pbom_version_gid: Optional[str] = Query(None, description="覆盖 BOP 版本绑定的 PBOM，用于 PBOM 升级后重新选择"),
    _u=Depends(_READ),
):
    """
    工位自动关联预览：
    1. 读取 BOP 版本的 pbom_version_gid（可被 query param 覆盖）
    2. 若 BOP 未绑定且无 query param，返回 need_select=True + 项目下可用 PBOM 版本列表
    3. 遍历 gbop_nav_bindings（confirmed=TRUE）→ 工序/操作/零件树
    4. 标记每条工序在本 BOP 版本中是否已创建过 bop_entry（linked）
    """
    with get_conn() as conn:
        with conn.cursor() as cur:
            # 1. 取 BOP 版本信息
            cur.execute(
                "SELECT project_gid, pbom_version_gid FROM workmanship_bop_bop_versions WHERE gid=%s",
                (bop_gid,)
            )
            bop_ver = cur.fetchone()
            if not bop_ver:
                raise HTTPException(404, f"BOP 版本 {bop_gid} 不存在")
            # query param 优先，其次取 BOP 已绑定值
            pbom_version_gid = pbom_version_gid or bop_ver['pbom_version_gid']
            project_gid = bop_ver['project_gid']

            if not pbom_version_gid:
                # 返回项目下可用 PBOM 版本列表供前端选择
                pbom_versions = []
                if project_gid:
                    cur.execute(
                        """
                        SELECT gid,
                               COALESCE(NULLIF(name,''), NULLIF(version_tag,''), gid) AS display_name,
                               status, created_at
                        FROM workmanship_bop_pbom_versions
                        WHERE project_gid=%s AND status='ready'
                        ORDER BY created_at DESC
                        """,
                        (project_gid,)
                    )
                    pbom_versions = [dict(r) for r in cur.fetchall()]
                return {"need_select": True, "pbom_versions": pbom_versions, "data": []}

            # 2. 取项目名
            project_name = ''
            if project_gid:
                cur.execute("SELECT name FROM workmanship_proj_projects WHERE gid=%s", (project_gid,))
                proj_row = cur.fetchone()
                if proj_row:
                    project_name = proj_row['name'] or ''

            # 3. 取 PBOM 版本名
            cur.execute(
                "SELECT COALESCE(NULLIF(name,''), NULLIF(version_tag,''), gid) AS display_name "
                "FROM workmanship_bop_pbom_versions WHERE gid=%s",
                (pbom_version_gid,)
            )
            pbom_row = cur.fetchone()
            pbom_version_name = pbom_row['display_name'] if pbom_row else pbom_version_gid

            # 4. 取该 PBOM 版本已确认的 gbop_nav_bindings（含 is_part_feed）
            cur.execute(
                """
                SELECT gbop_process_entry_gid, gbop_op_entry_gid, pbom_entry_gid,
                       COALESCE(is_part_feed, FALSE) AS is_part_feed
                FROM workmanship_bop_gbop_nav_bindings
                WHERE pbom_version_gid = %s AND confirmed = TRUE
                """,
                (pbom_version_gid,)
            )
            bindings = [dict(r) for r in cur.fetchall()]

            if not bindings:
                return {
                    "pbom_version": {
                        "gid": pbom_version_gid,
                        "name": pbom_version_name,
                        "project_name": project_name,
                    },
                    "data": [],
                }

            # 5. 收集所有涉及的 gbop_entry gid
            proc_gids = list({b['gbop_process_entry_gid'] for b in bindings if b['gbop_process_entry_gid']})
            op_gids   = list({b['gbop_op_entry_gid']      for b in bindings if b['gbop_op_entry_gid']})
            part_gids = list({b['pbom_entry_gid']          for b in bindings if b['pbom_entry_gid']})

            # 6. 取 GBOP 工序/操作 entry 信息
            all_entry_gids = proc_gids + op_gids
            entry_map: dict = {}
            if all_entry_gids:
                _ph = ",".join(["%s"] * len(all_entry_gids))
                cur.execute(
                    f"SELECT gid, vpps, vpps_desc, node_type, seq_no, parent_gid "
                    f"FROM workmanship_tpl_gbop_entries WHERE gid IN ({_ph})",
                    all_entry_gids,
                )
                for r in cur.fetchall():
                    entry_map[r['gid']] = dict(r)

            # 7. 取 PBOM 零件信息
            part_info_map: dict = {}
            if part_gids:
                _ph = ",".join(["%s"] * len(part_gids))
                cur.execute(
                    f"SELECT gid, vpps, title, part_no FROM workmanship_bop_pbom WHERE gid IN ({_ph})",
                    part_gids,
                )
                for r in cur.fetchall():
                    part_info_map[r['gid']] = dict(r)

            # 8. 取本 BOP 版本中已存在的 process 节点 vpps 集合（用于 linked 判定）
            cur.execute(
                "SELECT DISTINCT vpps FROM workmanship_bop_bop_entries "
                "WHERE version_gid=%s AND node_type='process' AND is_deleted=FALSE AND vpps IS NOT NULL",
                (bop_gid,)
            )
            existing_proc_vpps = {r['vpps'] for r in cur.fetchall()}

            # 8b. 取 station_process 节点，建立 proc_vpps → line_gids 映射（线体筛选用）
            cur.execute(
                """
                SELECT gid, parent_gid, child_vpps
                FROM workmanship_bop_bop_entries
                WHERE version_gid = %s AND node_type = 'station_process' AND is_deleted = FALSE
                """,
                (bop_gid,)
            )
            all_stations_preview = [dict(r) for r in cur.fetchall()]

            proc_vpps_to_line_gids: dict = {}
            raw_line_gids: set = set()
            for st in all_stations_preview:
                cvpps = st.get('child_vpps') or []
                if isinstance(cvpps, str):
                    cvpps = json.loads(cvpps)
                lgid = st.get('parent_gid')
                for cv in cvpps:
                    v = cv.get('vpps', '') if isinstance(cv, dict) else ''
                    if v:
                        proc_vpps_to_line_gids.setdefault(v, set()).add(lgid)
                        if lgid:
                            raw_line_gids.add(lgid)

            line_map_preview: dict = {}
            if raw_line_gids:
                _raw_list = list(raw_line_gids)
                _ph = ",".join(["%s"] * len(_raw_list))
                cur.execute(
                    f"SELECT gid, title, vpps, sort_order FROM workmanship_bop_bop_entries "
                    f"WHERE gid IN ({_ph}) AND is_deleted = FALSE",
                    _raw_list,
                )
                for r in cur.fetchall():
                    line_map_preview[r['gid']] = dict(r)

            # 各线体拥有多少个不同 proc_vpps
            line_proc_count: dict = {}
            for pv, lgids in proc_vpps_to_line_gids.items():
                for lgid in lgids:
                    if lgid:
                        line_proc_count.setdefault(lgid, set()).add(pv)

            lines = sorted([
                {
                    'gid':           lgid,
                    'title':         line_map_preview.get(lgid, {}).get('title') or lgid,
                    'vpps':          line_map_preview.get(lgid, {}).get('vpps') or '',
                    'process_count': len(pvs),
                }
                for lgid, pvs in line_proc_count.items()
            ], key=lambda x: line_map_preview.get(x['gid'], {}).get('sort_order') or 0)

    # 9. 组装扁平输出（parent_gid 供前端树形渲染）
    #    按 proc_entry_gid 分组 operations；按 op_entry_gid 分组 parts
    from collections import defaultdict
    op_by_proc: dict = defaultdict(list)
    parts_by_op: dict = defaultdict(dict)   # op_g -> {part_g: {info + is_part_feed}}

    for b in bindings:
        proc_g = b['gbop_process_entry_gid']
        op_g   = b['gbop_op_entry_gid']
        part_g = b['pbom_entry_gid']
        is_pf  = bool(b.get('is_part_feed', False))
        if op_g and proc_g:
            # 避免重复 op
            if op_g not in [x['gid'] for x in op_by_proc[proc_g]]:
                op_entry = entry_map.get(op_g, {})
                op_by_proc[proc_g].append({
                    'gid': op_g,
                    'vpps': op_entry.get('vpps', ''),
                    'title': op_entry.get('vpps_desc') or op_entry.get('vpps') or op_g,
                    'seq_no': op_entry.get('seq_no') or 0,
                })
        if op_g and part_g:
            part_info = part_info_map.get(part_g, {})
            existing = parts_by_op[op_g].get(part_g)
            if existing:
                # is_part_feed=TRUE 优先
                existing['is_part_feed'] = existing['is_part_feed'] or is_pf
            else:
                parts_by_op[op_g][part_g] = {
                    'gid':         part_g,
                    'vpps':        part_info.get('vpps', ''),
                    'title':       part_info.get('title') or part_info.get('part_no') or part_g,
                    'is_part_feed': is_pf,
                }

    data = []
    for proc_g in proc_gids:
        proc_entry = entry_map.get(proc_g, {})
        proc_vpps  = proc_entry.get('vpps', '')
        linked     = proc_vpps in existing_proc_vpps if proc_vpps else False

        data.append({
            'gid':       proc_g,
            'vpps':      proc_vpps,
            'title':     proc_entry.get('vpps_desc') or proc_vpps or proc_g,
            'type':      'process',
            'parent_gid': None,
            'linked':    linked,
            'seq_no':    proc_entry.get('seq_no') or 0,
            'line_gids': list(proc_vpps_to_line_gids.get(proc_vpps, set()) - {None}),
        })

        ops = sorted(op_by_proc.get(proc_g, []), key=lambda x: x['seq_no'])
        for op in ops:
            data.append({
                'gid':       op['gid'],
                'vpps':      op['vpps'],
                'title':     op['title'],
                'type':      'operation',
                'parent_gid': proc_g,
                'linked':    linked,
                'seq_no':    op['seq_no'],
            })
            for part in parts_by_op.get(op['gid'], {}).values():
                data.append({
                    'gid':         part['gid'],
                    'vpps':        part['vpps'],
                    'title':       part['title'],
                    'type':        'part',
                    'parent_gid':  op['gid'],
                    'linked':      linked,
                    'is_part_feed': part['is_part_feed'],
                })

    return {
        "pbom_version": {
            "gid": pbom_version_gid,
            "name": pbom_version_name,
            "project_name": project_name,
        },
        "lines": lines,
        "data": data,
    }


class StationAutolinkBody(BaseModel):
    pbom_version_gid: Optional[str] = None   # 覆盖/补填 BOP 版本绑定的 PBOM
    line_gids: Optional[List[str]] = None     # 仅处理指定线体下的工位；空=全部


@router.post("/bop-versions/{bop_gid}/station-autolink", status_code=200)
def station_autolink(bop_gid: str, body: StationAutolinkBody = StationAutolinkBody(), current_user: dict = Depends(_WRITE)):
    """
    工位自动关联执行：
    1. 取 BOP 版本 → pbom_version_gid（可由 body 覆盖，若覆盖则同时写入 bop_versions）
    2. 取已确认 gbop_nav_bindings 中的工序/操作/零件绑定
    3. 对每个 GBOP 工序（vpps=X）找 BOP 中 child_vpps 含 X 的 station_process 节点
    4. 幂等创建 process bop_entry（parent_gid=station.gid）
    5. 幂等创建 operation bop_entry + pbom_part bop_entry_links
    6. 更新 station 的 child_vpps 缓存
    """
    with get_conn() as conn:
        with conn.cursor() as cur:
            # 1. 取 BOP 版本信息
            cur.execute(
                "SELECT project_gid, pbom_version_gid FROM workmanship_bop_bop_versions WHERE gid=%s",
                (bop_gid,)
            )
            bop_ver = cur.fetchone()
            if not bop_ver:
                raise HTTPException(404, f"BOP 版本 {bop_gid} 不存在")
            pbom_version_gid = body.pbom_version_gid or bop_ver['pbom_version_gid']
            if not pbom_version_gid:
                raise HTTPException(400, "未提供 pbom_version_gid，请在面板中选择 PBOM 版本")
            # 若 BOP 版本尚未绑定，顺手写入
            if not bop_ver['pbom_version_gid'] and body.pbom_version_gid:
                cur.execute(
                    "UPDATE workmanship_bop_bop_versions SET pbom_version_gid=%s WHERE gid=%s",
                    (pbom_version_gid, bop_gid)
                )

            # 版本冻结检查
            cur.execute("SELECT frozen_at FROM workmanship_bop_bop_versions WHERE gid=%s", (bop_gid,))
            ver_row = cur.fetchone()
            if ver_row and ver_row['frozen_at']:
                raise HTTPException(403, "BOP 版本已冻结，不允许修改")

            # 2. 取已确认的 gbop_nav_bindings（含 is_part_feed 区分树节点/关联）
            cur.execute(
                """
                SELECT gbop_process_entry_gid, gbop_op_entry_gid, pbom_entry_gid,
                       COALESCE(is_part_feed, FALSE) AS is_part_feed
                FROM workmanship_bop_gbop_nav_bindings
                WHERE pbom_version_gid = %s AND confirmed = TRUE
                """,
                (pbom_version_gid,)
            )
            bindings = [dict(r) for r in cur.fetchall()]

            if not bindings:
                return {"ok": True, "created": 0, "skipped": 0, "message": "无已确认的绑定数据"}

            # 3. 取 GBOP 工序/操作 entry 信息
            proc_gids = list({b['gbop_process_entry_gid'] for b in bindings if b['gbop_process_entry_gid']})
            op_gids   = list({b['gbop_op_entry_gid']      for b in bindings if b['gbop_op_entry_gid']})
            part_gids = list({b['pbom_entry_gid']          for b in bindings if b['pbom_entry_gid']})

            entry_map: dict = {}
            if proc_gids + op_gids:
                _all_eg = proc_gids + op_gids
                _ph = ",".join(["%s"] * len(_all_eg))
                cur.execute(
                    f"SELECT gid, vpps, vpps_desc, node_type, seq_no, parent_gid "
                    f"FROM workmanship_tpl_gbop_entries WHERE gid IN ({_ph})",
                    _all_eg,
                )
                for r in cur.fetchall():
                    entry_map[r['gid']] = dict(r)

            # 3b. 取 PBOM 零件信息（title/vpps/part_no）
            part_info_map: dict = {}
            if part_gids:
                _ph = ",".join(["%s"] * len(part_gids))
                cur.execute(
                    f"SELECT gid, title, vpps, part_no FROM workmanship_bop_pbom WHERE gid IN ({_ph})",
                    part_gids,
                )
                for r in cur.fetchall():
                    part_info_map[r['gid']] = dict(r)

            # 4. 取 BOP 版本中 station_process 节点（含 child_vpps），可按线体过滤
            if body.line_gids:
                _ph = ",".join(["%s"] * len(body.line_gids))
                cur.execute(
                    f"""
                    SELECT gid, vpps, title, sort_order, level, child_vpps
                    FROM workmanship_bop_bop_entries
                    WHERE version_gid = %s AND node_type = 'station_process'
                      AND is_deleted = FALSE AND parent_gid IN ({_ph})
                    """,
                    [bop_gid] + list(body.line_gids),
                )
            else:
                cur.execute(
                    """
                    SELECT gid, vpps, title, sort_order, level, child_vpps
                    FROM workmanship_bop_bop_entries
                    WHERE version_gid = %s AND node_type = 'station_process' AND is_deleted = FALSE
                    """,
                    (bop_gid,)
                )
            stations = [dict(r) for r in cur.fetchall()]

            # 5. 建 vpps → 包含该 vpps 的工位列表
            #    station.child_vpps 是 [{vpps, node_type, title}, ...]
            vpps_to_stations: dict = {}
            for station in stations:
                cvpps = station.get('child_vpps') or []
                if isinstance(cvpps, str):
                    cvpps = json.loads(cvpps)
                for cv in cvpps:
                    v = cv.get('vpps', '') if isinstance(cv, dict) else ''
                    if v:
                        vpps_to_stations.setdefault(v, []).append(station)

            # 6. 按工序分组 operations 和 parts
            from collections import defaultdict
            ops_by_proc: dict = defaultdict(list)
            parts_by_op: dict = defaultdict(dict)   # op_g -> {part_gid: is_part_feed}
            for b in bindings:
                p_g  = b['gbop_process_entry_gid']
                op_g = b['gbop_op_entry_gid']
                pt_g = b['pbom_entry_gid']
                is_pf = bool(b.get('is_part_feed', False))
                if op_g and p_g and op_g not in ops_by_proc[p_g]:
                    ops_by_proc[p_g].append(op_g)
                if op_g and pt_g:
                    # is_part_feed=TRUE 优先
                    parts_by_op[op_g][pt_g] = parts_by_op[op_g].get(pt_g, False) or is_pf

            # 7. 逐工序写入
            created = 0
            skipped = 0
            stations_to_sync = set()

            # 7a. 清理旧版本生成的 operation 级 pbom_part is_primary 链接（避免与新树节点重复）
            cur.execute(
                """
                UPDATE workmanship_bop_bop_entry_links l
                SET is_deleted = TRUE
                FROM workmanship_bop_bop_entries e
                WHERE l.entry_gid = e.gid
                  AND e.version_gid = %s
                  AND e.node_type = 'operation'
                  AND l.link_type = 'pbom_part'
                  AND l.is_primary = TRUE
                  AND l.is_deleted = FALSE
                """,
                (bop_gid,)
            )

            for proc_g in proc_gids:
                proc_entry = entry_map.get(proc_g)
                if not proc_entry:
                    skipped += 1
                    continue
                proc_vpps  = proc_entry.get('vpps', '')
                if not proc_vpps:
                    skipped += 1
                    continue
                proc_title = proc_entry.get('vpps_desc') or proc_vpps

                matched_stations = vpps_to_stations.get(proc_vpps, [])
                if not matched_stations:
                    skipped += 1
                    continue

                for station in matched_stations:
                    station_gid   = station['gid']
                    station_level = station.get('level') or 2

                    # 幂等：检查工序是否已存在
                    cur.execute(
                        """
                        SELECT gid FROM workmanship_bop_bop_entries
                        WHERE version_gid=%s AND parent_gid=%s
                          AND node_type='process' AND vpps=%s AND is_deleted=FALSE
                        LIMIT 1
                        """,
                        (bop_gid, station_gid, proc_vpps)
                    )
                    existing_proc = cur.fetchone()
                    if existing_proc:
                        proc_entry_gid = existing_proc['gid']
                        skipped += 1
                    else:
                        # 创建工序 bop_entry
                        proc_entry_gid = str(next_gid())
                        cur.execute(
                            "INSERT INTO workmanship_bop_bop_entries"
                            "(gid, version_gid, parent_gid, node_type, sort_order, level, ai00_level,"
                            " title, vpps, vpps_desc, child_vpps, meta)"
                            " VALUES (%s,%s,%s,'process',%s,%s,4,%s,%s,%s,'[]','{}')",
                            (proc_entry_gid, bop_gid, station_gid,
                             proc_entry.get('seq_no') or 0,
                             station_level + 1,
                             proc_title, proc_vpps, proc_title)
                        )
                        created += 1
                        stations_to_sync.add(station_gid)

                    # 创建 operation entries
                    for op_g in ops_by_proc.get(proc_g, []):
                        op_entry = entry_map.get(op_g)
                        if not op_entry:
                            continue
                        op_vpps  = op_entry.get('vpps', '')
                        op_title = op_entry.get('vpps_desc') or op_vpps or op_g

                        # 幂等：检查操作是否已存在
                        cur.execute(
                            """
                            SELECT gid FROM workmanship_bop_bop_entries
                            WHERE version_gid=%s AND parent_gid=%s
                              AND node_type='operation' AND vpps=%s AND is_deleted=FALSE
                            LIMIT 1
                            """,
                            (bop_gid, proc_entry_gid, op_vpps) if op_vpps else (bop_gid, proc_entry_gid, op_g)
                        )
                        existing_op = cur.fetchone()
                        if existing_op:
                            op_entry_gid = existing_op['gid']
                        else:
                            op_entry_gid = str(next_gid())
                            cur.execute(
                                "INSERT INTO workmanship_bop_bop_entries"
                                "(gid, version_gid, parent_gid, node_type, sort_order, level, ai00_level,"
                                " title, vpps, vpps_desc, child_vpps, meta)"
                                " VALUES (%s,%s,%s,'operation',%s,%s,5,%s,%s,%s,'[]','{}')",
                                (op_entry_gid, bop_gid, proc_entry_gid,
                                 op_entry.get('seq_no') or 0,
                                 station_level + 2,
                                 op_title, op_vpps or None, op_title)
                            )
                            created += 1

                        # 创建 part 树节点（is_part_feed=True）或关联 link（is_part_feed=False）
                        for part_gid, is_part_feed in parts_by_op.get(op_g, {}).items():
                            part_info  = part_info_map.get(part_gid, {})
                            part_title = part_info.get('title') or part_info.get('part_no') or part_gid
                            part_vpps  = part_info.get('vpps')

                            if is_part_feed:
                                # 创建 bop_entry 树节点（node_type='part'），挂在 operation 下
                                if part_vpps:
                                    cur.execute(
                                        """
                                        SELECT gid FROM workmanship_bop_bop_entries
                                        WHERE version_gid=%s AND parent_gid=%s
                                          AND node_type='part' AND vpps=%s AND is_deleted=FALSE
                                        LIMIT 1
                                        """,
                                        (bop_gid, op_entry_gid, part_vpps)
                                    )
                                else:
                                    cur.execute(
                                        """
                                        SELECT e.gid FROM workmanship_bop_bop_entries e
                                        JOIN workmanship_bop_bop_entry_links l ON l.entry_gid = e.gid
                                        WHERE e.version_gid=%s AND e.parent_gid=%s
                                          AND e.node_type='part' AND l.link_type='pbom_part'
                                          AND l.entity_gid=%s AND l.is_deleted=FALSE
                                          AND e.is_deleted=FALSE
                                        LIMIT 1
                                        """,
                                        (bop_gid, op_entry_gid, part_gid)
                                    )
                                if not cur.fetchone():
                                    part_entry_gid = str(next_gid())
                                    cur.execute(
                                        "INSERT INTO workmanship_bop_bop_entries"
                                        "(gid, version_gid, parent_gid, node_type, sort_order,"
                                        " level, ai00_level, title, vpps, vpps_desc, child_vpps, meta)"
                                        " VALUES (%s,%s,%s,'part',0,%s,6,%s,%s,%s,'[]','{}')",
                                        (part_entry_gid, bop_gid, op_entry_gid,
                                         station_level + 3,
                                         part_title, part_vpps or None, part_title)
                                    )
                                    cur.execute(
                                        "INSERT INTO workmanship_bop_bop_entry_links"
                                        "(gid, entry_gid, version_gid, link_type, entity_gid,"
                                        " is_primary, is_inherited)"
                                        " VALUES (%s,%s,%s,'pbom_part',%s,TRUE,FALSE)",
                                        (str(next_gid()), part_entry_gid, bop_gid, part_gid)
                                    )
                                    created += 1
                            else:
                                # 非 part_feed：在 operation 节点挂关联 link（is_primary=FALSE）
                                cur.execute(
                                    """
                                    SELECT gid FROM workmanship_bop_bop_entry_links
                                    WHERE entry_gid=%s AND link_type='pbom_part'
                                      AND entity_gid=%s AND is_deleted=FALSE
                                    LIMIT 1
                                    """,
                                    (op_entry_gid, part_gid)
                                )
                                if not cur.fetchone():
                                    cur.execute(
                                        "INSERT INTO workmanship_bop_bop_entry_links"
                                        "(gid, entry_gid, version_gid, link_type, entity_gid,"
                                        " is_primary, is_inherited)"
                                        " VALUES (%s,%s,%s,'pbom_part',%s,FALSE,FALSE)",
                                        (str(next_gid()), op_entry_gid, bop_gid, part_gid)
                                    )

            # 8. 更新 station 的 child_vpps 缓存
            for station_gid in stations_to_sync:
                cur.execute(
                    """
                    SELECT e.gid, e.node_type, e.title, e.vpps
                    FROM workmanship_bop_bop_entries e
                    WHERE e.parent_gid = %s AND e.version_gid = %s AND e.is_deleted = FALSE
                      AND e.vpps IS NOT NULL
                    """,
                    (station_gid, bop_gid)
                )
                children = cur.fetchall()
                child_vpps_arr = [
                    {"vpps": c['vpps'], "node_type": c['node_type'], "title": c['title'] or ''}
                    for c in children
                ]
                cur.execute(
                    "UPDATE workmanship_bop_bop_entries SET child_vpps=%s, updated_at=NOW() WHERE gid=%s",
                    (json.dumps(child_vpps_arr), station_gid)
                )

            conn.commit()

    return {"ok": True, "created": created, "skipped": skipped}


@router.post("/bop-versions/{bop_gid}/station-autolink-undo", status_code=200)
def station_autolink_undo(
    bop_gid: str,
    mode: str = Query("soft", description="soft=软删除（默认）；hard=硬删除（超管专属）"),
    _u=Depends(_WRITE),
):
    """
    撤销工位自动关联：删除本 BOP 版本中由 auto-link 创建的
    process / operation / part bop_entries 及其 bop_entry_links。
    mode=soft（默认）：软删除；mode=hard：硬删除，需超管权限。
    """
    if mode == "hard":
        # 硬删除需超管权限
        user_role = (_u.get("role") or _u.get("system_role") or "") if isinstance(_u, dict) else ""
        if user_role != "super_admin":
            raise HTTPException(403, "硬删除需要超管权限")
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT gid, frozen_at FROM workmanship_bop_bop_versions WHERE gid=%s",
                (bop_gid,)
            )
            ver = cur.fetchone()
            if not ver:
                raise HTTPException(404, f"BOP 版本 {bop_gid} 不存在")
            if ver['frozen_at']:
                raise HTTPException(403, "BOP 版本已冻结，不允许修改")

            # 找 station_process 节点
            cur.execute(
                """
                SELECT gid FROM workmanship_bop_bop_entries
                WHERE version_gid = %s AND node_type = 'station_process' AND is_deleted = FALSE
                """,
                (bop_gid,)
            )
            station_gids = [r['gid'] for r in cur.fetchall()]
            if not station_gids:
                return {"ok": True, "deleted": 0}

            # 找 process 子节点
            _ph = ",".join(["%s"] * len(station_gids))
            cur.execute(
                f"""
                SELECT gid FROM workmanship_bop_bop_entries
                WHERE version_gid = %s AND node_type = 'process'
                  AND parent_gid IN ({_ph}) AND is_deleted = FALSE
                """,
                [bop_gid] + station_gids,
            )
            proc_gids = [r['gid'] for r in cur.fetchall()]

            # 找 operation 子节点
            op_gids = []
            if proc_gids:
                _ph = ",".join(["%s"] * len(proc_gids))
                cur.execute(
                    f"""
                    SELECT gid FROM workmanship_bop_bop_entries
                    WHERE version_gid = %s AND node_type = 'operation'
                      AND parent_gid IN ({_ph}) AND is_deleted = FALSE
                    """,
                    [bop_gid] + proc_gids,
                )
                op_gids = [r['gid'] for r in cur.fetchall()]

            # 找 part 子节点
            part_gids = []
            if op_gids:
                _ph = ",".join(["%s"] * len(op_gids))
                cur.execute(
                    f"""
                    SELECT gid FROM workmanship_bop_bop_entries
                    WHERE version_gid = %s AND node_type = 'part'
                      AND parent_gid IN ({_ph}) AND is_deleted = FALSE
                    """,
                    [bop_gid] + op_gids,
                )
                part_gids = [r['gid'] for r in cur.fetchall()]

            all_gids = proc_gids + op_gids + part_gids
            if not all_gids:
                return {"ok": True, "deleted": 0}

            if mode == "hard":
                # 硬删除：直接 DELETE
                _ph = ",".join(["%s"] * len(all_gids))
                cur.execute(
                    f"DELETE FROM workmanship_bop_bop_entry_links WHERE entry_gid IN ({_ph})",
                    all_gids,
                )
                cur.execute(
                    f"DELETE FROM workmanship_bop_bop_entries WHERE gid IN ({_ph})",
                    all_gids,
                )
            else:
                # 软删除（默认）
                _ph = ",".join(["%s"] * len(all_gids))
                cur.execute(
                    f"""
                    UPDATE workmanship_bop_bop_entry_links SET is_deleted = TRUE
                    WHERE entry_gid IN ({_ph}) AND is_deleted = FALSE
                    """,
                    all_gids,
                )
                cur.execute(
                    f"""
                    UPDATE workmanship_bop_bop_entries SET is_deleted = TRUE, updated_at = NOW()
                    WHERE gid IN ({_ph})
                    """,
                    all_gids,
                )
            conn.commit()

    return {"ok": True, "deleted": len(all_gids)}