"""
backend/routers/import_export.py
──────────────────────────────────
通用导入导出 API

端点：
  GET    /api/import-export/templates              列出模板（本人 + shared）
  POST   /api/import-export/templates              创建模板
  PATCH  /api/import-export/templates/{gid}        更新模板
  DELETE /api/import-export/templates/{gid}        删除模板
  POST   /api/import-export/export/excel           依据 template_config + rows 生成 .xlsx
  POST   /api/import-export/import/parse-excel     接收 base64 Excel → 返回预览数据
  POST   /api/import-export/lark-sheets/read       飞书电子表格读取
  POST   /api/import-export/lark-sheets/write      飞书电子表格写入
  POST   /api/import-export/lark-bitable/read      飞书多维表读取
  POST   /api/import-export/lark-bitable/write     飞书多维表写入
"""
import base64
import io
import json
import re
from datetime import datetime
from typing import Any, Dict, List, Optional

import httpx
from fastapi import APIRouter, Depends, HTTPException, Request
from pydantic import BaseModel

from ..data.connection import get_conn
from backend.platform_sdk.auth import require_role
from backend.platform_sdk.ids import next_gid
from backend.capability_v2.gateway import get_default_gateway
from backend.platform_sdk.auth import get_authenticated_principal
from backend.platform_sdk.factory import build_web_compatibility_envelope, invoke_compatibility
from uuid import uuid4

router = APIRouter(prefix="/api/import-export", tags=["import-export"])

_READ = require_role(
    "super_admin", "team_admin", "project_admin",
    "rule_admin", "knowledge_admin", "member",
)

_TMPL_KEYS = ["gid", "name", "module", "owner_gid", "is_shared",
              "config", "created_at", "updated_at"]


# ── Pydantic 模型 ─────────────────────────────────────────────────────────────

class CreateTemplateBody(BaseModel):
    name: str = "未命名模板"
    module: str = "*"
    config: dict = {}
    is_shared: bool = False


class UpdateTemplateBody(BaseModel):
    name: Optional[str] = None
    module: Optional[str] = None
    config: Optional[dict] = None
    is_shared: Optional[bool] = None


class ExportExcelBody(BaseModel):
    template_config: dict
    rows: List[Dict[str, Any]]
    filename: str = ""


class ParseExcelBody(BaseModel):
    file_b64: str
    filename: str = "upload.xlsx"


class LarkSheetsReadBody(BaseModel):
    user_access_token: str
    spreadsheet_token: str
    sheet_range: str = "Sheet1!A1:Z1000"


class LarkSheetsWriteBody(BaseModel):
    user_access_token: str
    spreadsheet_token: str
    sheet_id: str
    headers: List[str]
    rows: List[List[Any]]


class LarkBitableReadBody(BaseModel):
    user_access_token: str
    app_token: str
    table_id: str
    page_size: int = 500


class LarkBitableWriteBody(BaseModel):
    user_access_token: str
    app_token: str
    table_id: str
    records: List[Dict[str, Any]]


# ── 辅助函数 ──────────────────────────────────────────────────────────────────

def _row_to_dict(row, keys):
    return dict(zip(keys, row))


# ── 模板 CRUD ─────────────────────────────────────────────────────────────────

@router.get("/templates")
async def list_templates(
    request: Request,
    module: str = "",
    user=Depends(_READ),
    principal=Depends(get_authenticated_principal),
    gateway=Depends(get_default_gateway),
):
    request_id = request.headers.get("X-Request-ID") or f"base_export_template_{uuid4().hex}"
    result = await invoke_compatibility(
        gateway,
        build_web_compatibility_envelope(
            gateway,
            capability_id="base.export_template.read",
            payload={"module": module, "limit": 500},
            current_user=user,
            principal=principal,
            request_id=request_id,
            trace_id=request.headers.get("X-Trace-ID") or request_id,
        ),
    )
    if not result.ok:
        error = result.error
        code = error.code if error else "provider_error"
        raise HTTPException(
            status_code={
                "resource_not_found": 404,
                "permission_denied": 403,
                "invalid_input": 400,
                "response_limit_exceeded": 413,
            }.get(code, 422),
            detail=error.model_dump(mode="json") if error else None,
        )
    return {"success": True, "data": result.data["data"]["items"]}


async def _invoke_template_change(request, user, principal, gateway, payload):
    request_id = request.headers.get("X-Request-ID") or f"base_export_template_change_{uuid4().hex}"
    result = await invoke_compatibility(
        gateway,
        build_web_compatibility_envelope(
            gateway,
            capability_id="base.export_template.change.apply",
            payload=payload,
            current_user=user,
            principal=principal,
            request_id=request_id,
            trace_id=request.headers.get("X-Trace-ID") or request_id,
            idempotency_key=request.headers.get("X-Idempotency-Key") or request_id,
            approval_reference=request.headers.get("X-Capability-Approval"),
        ),
    )
    if not result.ok:
        error = result.error
        code = error.code if error else "provider_error"
        raise HTTPException(
            status_code={
                "resource_not_found": 404,
                "permission_denied": 403,
                "invalid_input": 400,
                "confirmation_required": 409,
            }.get(code, 422),
            detail=error.model_dump(mode="json") if error else None,
        )
    return result.data["data"]


async def _invoke_data_exchange(request, user, principal, gateway, payload):
    request_id = request.headers.get("X-Request-ID") or f"craft_data_exchange_{uuid4().hex}"
    result = await invoke_compatibility(
        gateway,
        build_web_compatibility_envelope(
            gateway,
            capability_id="craft.data_exchange.export",
            payload=payload,
            current_user=user,
            principal=principal,
            request_id=request_id,
            trace_id=request.headers.get("X-Trace-ID") or request_id,
            idempotency_key=request.headers.get("X-Idempotency-Key") or request_id,
            approval_reference=request.headers.get("X-Capability-Approval"),
        ),
    )
    if not result.ok:
        error = result.error
        code = error.code if error else "provider_error"
        raise HTTPException(
            status_code={"invalid_input": 400, "confirmation_required": 409}.get(code, 422),
            detail=error.model_dump(mode="json") if error else None,
        )
    return result.data["data"]


async def _invoke_lark_read(request, user, principal, gateway, payload):
    request_id = request.headers.get("X-Request-ID") or f"craft_lark_read_{uuid4().hex}"
    result = await invoke_compatibility(gateway, build_web_compatibility_envelope(
        gateway, capability_id="craft.data_exchange.lark.read", payload=payload,
        current_user=user, principal=principal, request_id=request_id,
        trace_id=request.headers.get("X-Trace-ID") or request_id,
    ))
    if not result.ok:
        error = result.error
        code = error.code if error else "provider_error"
        raise HTTPException(status_code={"invalid_input": 400, "permission_denied": 403, "response_limit_exceeded": 413}.get(code, 422), detail=error.model_dump(mode="json") if error else None)
    return result.data["data"]


async def _invoke_lark_write(request, user, principal, gateway, payload):
    request_id = request.headers.get("X-Request-ID") or f"craft_lark_write_{uuid4().hex}"
    result = await invoke_compatibility(gateway, build_web_compatibility_envelope(
        gateway, capability_id="craft.data_exchange.lark.write", payload=payload,
        current_user=user, principal=principal, request_id=request_id,
        trace_id=request.headers.get("X-Trace-ID") or request_id,
        idempotency_key=request.headers.get("X-Idempotency-Key") or request_id,
        approval_reference=request.headers.get("X-Capability-Approval"),
    ))
    if not result.ok:
        error = result.error
        code = error.code if error else "provider_error"
        raise HTTPException(status_code={"invalid_input": 400, "permission_denied": 403, "confirmation_required": 409}.get(code, 422), detail=error.model_dump(mode="json") if error else None)
    return result.data["data"]


@router.post("/templates")
async def create_template(
    body: CreateTemplateBody,
    request: Request,
    user=Depends(_READ),
    principal=Depends(get_authenticated_principal),
    gateway=Depends(get_default_gateway),
):
    data = await _invoke_template_change(
        request,
        user,
        principal,
        gateway,
        {
            "operation": "create",
            "name": body.name,
            "module": body.module,
            "config": body.config,
            "is_shared": body.is_shared,
        },
    )
    return {"success": True, "data": data}


def _template_admin(user: dict) -> bool:
    role = user.get("org_role") or user.get("system_role") or "external"
    return role in {"super_admin", "team_admin"} or user.get("system_role") == "team_admin"


@router.patch("/templates/{gid}")
async def update_template(
    gid: str,
    body: UpdateTemplateBody,
    request: Request,
    user=Depends(_READ),
    principal=Depends(get_authenticated_principal),
    gateway=Depends(get_default_gateway),
):
    data = await _invoke_template_change(
        request,
        user,
        principal,
        gateway,
        {
            "operation": "update",
            "gid": gid,
            "updates": body.model_dump(exclude_none=True),
        },
    )
    return {"success": True, "data": data}


@router.delete("/templates/{gid}")
async def delete_template(
    gid: str,
    request: Request,
    user=Depends(_READ),
    principal=Depends(get_authenticated_principal),
    gateway=Depends(get_default_gateway),
):
    data = await _invoke_template_change(
        request,
        user,
        principal,
        gateway,
        {"operation": "delete", "gid": gid},
    )
    return {"success": True, "data": data}


# ── Excel 导出 ────────────────────────────────────────────────────────────────

@router.post("/export/excel")
async def export_excel(
    body: ExportExcelBody,
    request: Request,
    user=Depends(_READ),
    principal=Depends(get_authenticated_principal),
    gateway=Depends(get_default_gateway),
):
    data = await _invoke_data_exchange(
        request, user, principal, gateway,
        {"operation": "excel", "template_config": body.template_config, "rows": body.rows, "filename": body.filename},
    )
    return {"success": True, "data": data}


def _legacy_export_excel(body: ExportExcelBody, user=Depends(_READ)):
    """依据 template_config + rows 生成带样式 .xlsx，返回 base64。"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(500, "服务端缺少 openpyxl 依赖，请运行 pip install openpyxl")

    cfg = body.template_config
    columns = [c for c in cfg.get("columns", []) if c.get("include", True)]
    styles = cfg.get("styles", {})

    header_bg = styles.get("headerBg", "#2563EB").lstrip("#")
    header_fg = styles.get("headerFg", "#FFFFFF").lstrip("#")
    alt_row_bg = styles.get("altRowBg", "#EFF6FF").lstrip("#")
    font_size = int(styles.get("fontSize", 11))
    border_style = styles.get("borderStyle", "thin")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "导出数据"

    # 表头样式
    hdr_fill = PatternFill("solid", fgColor=header_bg)
    hdr_font = Font(bold=True, color=header_fg, size=font_size)
    alt_fill = PatternFill("solid", fgColor=alt_row_bg)
    thin_border = Border(
        left=Side(style=border_style),
        right=Side(style=border_style),
        top=Side(style=border_style),
        bottom=Side(style=border_style),
    )
    center_align = Alignment(horizontal="center", vertical="center")

    # 写表头
    for col_idx, col in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col.get("label", col.get("key", "")))
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.border = thin_border
        cell.alignment = center_align
        ws.column_dimensions[get_column_letter(col_idx)].width = int(col.get("width", 15))

    ws.row_dimensions[1].height = 20

    # 写数据行
    for row_idx, row in enumerate(body.rows, 2):
        is_alt = (row_idx % 2 == 0)
        for col_idx, col in enumerate(columns, 1):
            val = row.get(col["key"], "")
            # 复杂值转字符串
            if isinstance(val, (dict, list)):
                val = json.dumps(val, ensure_ascii=False)
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            if is_alt:
                cell.fill = alt_fill

    # 生成 base64
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    file_b64 = base64.b64encode(buf.read()).decode()

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = body.filename or f"export_{ts}.xlsx"

    return {"success": True, "data": {"file_b64": file_b64, "filename": filename}}


# ── Excel 解析（导入） ────────────────────────────────────────────────────────

@router.post("/import/parse-excel")
async def parse_excel(
    body: ParseExcelBody,
    request: Request,
    user=Depends(_READ),
    principal=Depends(get_authenticated_principal),
    gateway=Depends(get_default_gateway),
):
    parsed = _legacy_parse_excel(body, user)
    parsed_data = parsed.get("data", {})
    document = {
        "version_tag": body.filename.rsplit(".", 1)[0] or "import",
        "bop_name": body.filename,
        "entries": [
            dict(zip(parsed_data.get("headers", []), row))
            for row in parsed_data.get("rows", [])
        ],
    }
    request_id = request.headers.get("X-Request-ID") or f"craft_bop_import_preview_{uuid4().hex}"
    result = await invoke_compatibility(
        gateway,
        build_web_compatibility_envelope(
            gateway,
            capability_id="craft.bop.import.preview",
            payload={"document": document},
            current_user=user,
            principal=principal,
            request_id=request_id,
            trace_id=request.headers.get("X-Trace-ID") or request_id,
        ),
    )
    if not result.ok:
        error = result.error
        raise HTTPException(
            status_code=400 if error and error.code == "invalid_input" else 422,
            detail=error.model_dump(mode="json") if error else None,
        )
    parsed_data["import_preview"] = result.data["data"]
    return parsed


def _legacy_parse_excel(body: ParseExcelBody, user=Depends(_READ)):
    """接收 base64 编码的 Excel / CSV 文件，返回 {headers, rows, warnings}。"""
    try:
        file_bytes = base64.b64decode(body.file_b64)
    except Exception:
        raise HTTPException(400, "file_b64 解码失败")

    is_csv = body.filename.lower().endswith('.csv')

    if is_csv:
        # ── CSV 分支 ──────────────────────────────────────────────────────────
        import csv as _csv

        # 尝试常见编码：UTF-8-BOM → UTF-8 → GBK → Latin-1
        text = None
        for enc in ('utf-8-sig', 'utf-8', 'gbk', 'latin-1'):
            try:
                text = file_bytes.decode(enc)
                break
            except (UnicodeDecodeError, LookupError):
                continue
        if text is None:
            text = file_bytes.decode('latin-1', errors='replace')

        reader = _csv.reader(text.splitlines())
        all_rows = list(reader)
        if not all_rows:
            return {"success": True, "data": {"headers": [], "rows": [], "warnings": ["文件为空"]}}

        # 跳过前导空行找表头
        header_idx = 0
        for idx, raw_row in enumerate(all_rows):
            if any(str(v).strip() for v in raw_row):
                header_idx = idx
                break

        headers = [str(h).strip() or f"列{i+1}" for i, h in enumerate(all_rows[header_idx])]
        col_count = len(headers)

        warnings = []
        rows = []
        for raw_row in all_rows[header_idx + 1:]:
            padded = list(raw_row) + [''] * max(0, col_count - len(raw_row))
            trimmed = padded[:col_count]
            if all(str(v).strip() == '' for v in trimmed):
                continue
            rows.append([str(v).strip() for v in trimmed])

        if not rows:
            warnings.append("未找到有效数据行（表头之外无数据）")
        return {"success": True, "data": {"headers": headers, "rows": rows, "warnings": warnings}}

    # ── Excel 分支 ────────────────────────────────────────────────────────────
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(500, "服务端缺少 openpyxl 依赖，请运行 pip install openpyxl")

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, keep_vba=False, data_only=True)
        ws = wb.active

        warnings = []
        all_rows = list(ws.iter_rows(values_only=True))

        if not all_rows:
            return {"success": True, "data": {"headers": [], "rows": [], "warnings": ["文件为空"]}}

        # 找到第一个有数据的行作为表头（跳过前导空行）
        header_idx = 0
        for idx, raw_row in enumerate(all_rows):
            if any(v is not None and str(v).strip() != "" for v in raw_row):
                header_idx = idx
                break

        raw_headers = all_rows[header_idx]
        # 找到最后一个有值的列位置，截掉右侧空列
        last_col = 0
        for i, h in enumerate(raw_headers):
            if h is not None and str(h).strip() != "":
                last_col = i
        col_count = last_col + 1

        headers = [str(h).strip() if h is not None else f"列{i+1}" for i, h in enumerate(raw_headers[:col_count])]

        rows = []
        for i, raw_row in enumerate(all_rows[header_idx + 1:], start=header_idx + 2):
            # 截到 col_count 列
            trimmed = raw_row[:col_count] if len(raw_row) >= col_count else raw_row
            # 跳过全空行
            if all(v is None or str(v).strip() == "" for v in trimmed):
                continue
            rows.append([str(v).strip() if v is not None else "" for v in trimmed])

        if not rows:
            warnings.append("未找到有效数据行（表头之外无数据）")

        return {"success": True, "data": {"headers": headers, "rows": rows, "warnings": warnings}}
    except Exception as e:
        raise HTTPException(400, f"Excel 解析失败：{str(e)}")


# ── 飞书电子表格 ──────────────────────────────────────────────────────────────

@router.post("/lark-sheets/read")
async def lark_sheets_read(body: LarkSheetsReadBody, request: Request, user=Depends(_READ), principal=Depends(get_authenticated_principal), gateway=Depends(get_default_gateway)):
    return await _invoke_lark_read(request, user, principal, gateway, {"operation": "sheets.read", **body.model_dump()})


async def _legacy_lark_sheets_read(body: LarkSheetsReadBody, user=Depends(_READ)):
    """读取飞书电子表格数据，返回 {headers, rows}。"""
    url = (
        f"https://open.feishu.cn/open-apis/sheets/v2/spreadsheets"
        f"/{body.spreadsheet_token}/values/{body.sheet_range}"
    )
    async with httpx.AsyncClient(timeout=30) as client:
        resp = await client.get(
            url,
            headers={"Authorization": f"Bearer {body.user_access_token}"},
        )
    if resp.status_code != 200:
        raise HTTPException(resp.status_code, f"飞书 API 错误：{resp.text}")
    data = resp.json()
    if data.get("code") != 0:
        raise HTTPException(400, f"飞书返回错误：{data.get('msg')}")

    value_range = data.get("data", {}).get("valueRange", {})
    values = value_range.get("values", [])
    if not values:
        return {"success": True, "data": {"headers": [], "rows": [], "warnings": ["表格为空"]}}

    headers = [str(v) if v is not None else f"列{i+1}" for i, v in enumerate(values[0])]
    rows = [[str(v) if v is not None else "" for v in row] for row in values[1:]]
    return {"success": True, "data": {"headers": headers, "rows": rows, "warnings": []}}


@router.post("/lark-sheets/write")
async def lark_sheets_write(body: LarkSheetsWriteBody, request: Request, user=Depends(_READ), principal=Depends(get_authenticated_principal), gateway=Depends(get_default_gateway)):
    return await _invoke_lark_write(request, user, principal, gateway, {"operation": "sheets.write", "user_access_token": body.user_access_token, "spreadsheet_token": body.spreadsheet_token, "sheet_id": body.sheet_id, "values": [body.headers] + body.rows})


async def _legacy_lark_sheets_write(body: LarkSheetsWriteBody, user=Depends(_READ)):
    """向飞书电子表格写入数据（覆盖指定 sheet）。"""
    values = [body.headers] + body.rows
    url = (
        f"https://open.feishu.cn/open-apis/sheets/v2/spreadsheets"
        f"/{body.spreadsheet_token}/values"
    )
    payload = {
        "valueRange": {
            "range": f"{body.sheet_id}!A1:{_col_letter(len(body.headers))}{len(values)}",
            "values": values,
        }
    }
    async with httpx.AsyncClient(timeout=30) as client:
        resp = await client.put(
            url,
            headers={
                "Authorization": f"Bearer {body.user_access_token}",
                "Content-Type": "application/json",
            },
            json=payload,
        )
    if resp.status_code != 200:
        raise HTTPException(resp.status_code, f"飞书 API 错误：{resp.text}")
    data = resp.json()
    if data.get("code") != 0:
        raise HTTPException(400, f"飞书返回错误：{data.get('msg')}")
    return {"success": True, "data": {"written_rows": len(body.rows)}}


# ── 飞书多维表 ────────────────────────────────────────────────────────────────

@router.post("/lark-bitable/read")
async def lark_bitable_read(body: LarkBitableReadBody, request: Request, user=Depends(_READ), principal=Depends(get_authenticated_principal), gateway=Depends(get_default_gateway)):
    return await _invoke_lark_read(request, user, principal, gateway, {"operation": "bitable.read", **body.model_dump()})


async def _legacy_lark_bitable_read(body: LarkBitableReadBody, user=Depends(_READ)):
    """读取飞书多维表所有记录，返回 {headers, rows}。"""
    base_url = (
        f"https://open.feishu.cn/open-apis/bitable/v1/apps"
        f"/{body.app_token}/tables/{body.table_id}/records"
    )
    headers_auth = {"Authorization": f"Bearer {body.user_access_token}"}
    all_records = []
    page_token = None

    async with httpx.AsyncClient(timeout=30) as client:
        while True:
            params: dict = {"page_size": body.page_size}
            if page_token:
                params["page_token"] = page_token
            resp = await client.get(base_url, headers=headers_auth, params=params)
            if resp.status_code != 200:
                raise HTTPException(resp.status_code, f"飞书 API 错误：{resp.text}")
            data = resp.json()
            if data.get("code") != 0:
                raise HTTPException(400, f"飞书返回错误：{data.get('msg')}")
            items = data.get("data", {}).get("items", [])
            all_records.extend(items)
            if not data.get("data", {}).get("has_more"):
                break
            page_token = data["data"].get("page_token")

    if not all_records:
        return {"success": True, "data": {"headers": [], "rows": [], "warnings": ["多维表为空"]}}

    # 从第一条记录收集字段名
    field_names = list(all_records[0].get("fields", {}).keys())
    rows = []
    for rec in all_records:
        fields = rec.get("fields", {})
        rows.append([_bitable_val(fields.get(f)) for f in field_names])

    return {"success": True, "data": {"headers": field_names, "rows": rows, "warnings": []}}


@router.post("/lark-bitable/write")
async def lark_bitable_write(body: LarkBitableWriteBody, request: Request, user=Depends(_READ), principal=Depends(get_authenticated_principal), gateway=Depends(get_default_gateway)):
    return await _invoke_lark_write(request, user, principal, gateway, {"operation": "bitable.write", **body.model_dump()})


async def _legacy_lark_bitable_write(body: LarkBitableWriteBody, user=Depends(_READ)):
    """批量写入飞书多维表记录。"""
    url = (
        f"https://open.feishu.cn/open-apis/bitable/v1/apps"
        f"/{body.app_token}/tables/{body.table_id}/records/batch_create"
    )
    # 每次最多写 500 条
    written = 0
    async with httpx.AsyncClient(timeout=60) as client:
        for i in range(0, len(body.records), 500):
            chunk = body.records[i:i+500]
            payload = {"records": [{"fields": r} for r in chunk]}
            resp = await client.post(
                url,
                headers={
                    "Authorization": f"Bearer {body.user_access_token}",
                    "Content-Type": "application/json",
                },
                json=payload,
            )
            if resp.status_code != 200:
                raise HTTPException(resp.status_code, f"飞书 API 错误：{resp.text}")
            data = resp.json()
            if data.get("code") != 0:
                raise HTTPException(400, f"飞书返回错误：{data.get('msg')}")
            written += len(chunk)

    return {"success": True, "data": {"written_records": written}}


# ── 工具函数 ──────────────────────────────────────────────────────────────────

def _col_letter(n: int) -> str:
    """列数转 Excel 列字母（1→A, 27→AA…）。"""
    result = ""
    while n > 0:
        n, rem = divmod(n - 1, 26)
        result = chr(65 + rem) + result
    return result or "A"


def _bitable_val(v) -> str:
    """多维表字段值转字符串（处理嵌套结构）。"""
    if v is None:
        return ""
    if isinstance(v, list):
        # 富文本/人员等类型
        parts = []
        for item in v:
            if isinstance(item, dict):
                parts.append(item.get("text") or item.get("name") or str(item))
            else:
                parts.append(str(item))
        return ", ".join(parts)
    return str(v)


# ── Diff 报告导出 ──────────────────────────────────────────────────────────────

class DiffReportBody(BaseModel):
    columns:   List[dict]       # [{key, label, width}]
    diff_rows: List[dict]       # [{_diffStatus, _rowA, _rowB, _changedFields}]
    label_a:   str = "数据集 A"
    label_b:   str = "数据集 B"
    filename:  str = ""


@router.post("/export/diff-report")
async def export_diff_report(
    body: DiffReportBody,
    request: Request,
    user=Depends(_READ),
    principal=Depends(get_authenticated_principal),
    gateway=Depends(get_default_gateway),
):
    data = await _invoke_data_exchange(
        request, user, principal, gateway,
        {
            "operation": "diff_report", "columns": body.columns, "diff_rows": body.diff_rows,
            "label_a": body.label_a, "label_b": body.label_b, "filename": body.filename,
        },
    )
    return {"success": True, "data": data}


def _legacy_export_diff_report(body: DiffReportBody, user=Depends(_READ)):
    """生成差异对比 .xlsx（A侧列 + B侧列，行按 diffStatus 着色）。"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(500, "服务端缺少 openpyxl 依赖，请运行 pip install openpyxl")

    # 颜色定义（去掉 # 前缀）
    COLOR_ADDED    = "DCFCE7"   # 淡绿
    COLOR_REMOVED  = "FEE2E2"   # 淡红
    COLOR_MODIFIED = "FEF3C7"   # 淡橙
    COLOR_CHANGED_CELL = "F9E2AF"  # 变更字段单元格（更深橙）
    COLOR_HEADER_A = "1E3A5F"   # 深蓝（A侧标题行）
    COLOR_HEADER_B = "1E4D2B"   # 深绿（B侧标题行）
    COLOR_COL_HDR  = "374151"   # 列名行背景

    cols = body.columns
    n = len(cols)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "差异对比"

    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # ── 第1行：A侧/B侧合并标题 ──────────────────────────────────────────────
    ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=n + 1)
    ws.merge_cells(start_row=1, start_column=n + 2, end_row=1, end_column=2 * n + 1)

    cell_a = ws.cell(row=1, column=2, value=body.label_a)
    cell_a.fill = PatternFill("solid", fgColor=COLOR_HEADER_A)
    cell_a.font = Font(bold=True, color="FFFFFF", size=12)
    cell_a.alignment = center

    cell_b = ws.cell(row=1, column=n + 2, value=body.label_b)
    cell_b.fill = PatternFill("solid", fgColor=COLOR_HEADER_B)
    cell_b.font = Font(bold=True, color="FFFFFF", size=12)
    cell_b.alignment = center

    # 状态列标题
    status_cell = ws.cell(row=1, column=1, value="状态")
    status_cell.fill = PatternFill("solid", fgColor=COLOR_COL_HDR)
    status_cell.font = Font(bold=True, color="FFFFFF", size=11)
    status_cell.alignment = center

    # ── 第2行：列名表头 ──────────────────────────────────────────────────────
    ws.cell(row=2, column=1, value="").fill = PatternFill("solid", fgColor=COLOR_COL_HDR)
    for i, col in enumerate(cols):
        label = col.get("label", col.get("key", ""))
        col_width = int(col.get("width", 12))
        # A侧
        ca = ws.cell(row=2, column=2 + i, value=label)
        ca.fill = PatternFill("solid", fgColor=COLOR_HEADER_A)
        ca.font = Font(bold=True, color="FFFFFF", size=10)
        ca.alignment = center
        ca.border = thin
        ws.column_dimensions[get_column_letter(2 + i)].width = max(8, min(col_width, 30))
        # B侧
        cb = ws.cell(row=2, column=n + 2 + i, value=label)
        cb.fill = PatternFill("solid", fgColor=COLOR_HEADER_B)
        cb.font = Font(bold=True, color="FFFFFF", size=10)
        cb.alignment = center
        cb.border = thin
        ws.column_dimensions[get_column_letter(n + 2 + i)].width = max(8, min(col_width, 30))

    ws.column_dimensions["A"].width = 8
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 20

    STATUS_MAP = {
        "added":    ("＋新增", COLOR_ADDED),
        "removed":  ("－删除", COLOR_REMOVED),
        "modified": ("～变更", COLOR_MODIFIED),
        "same":     ("＝相同", "FFFFFF"),
    }

    # ── 数据行 ────────────────────────────────────────────────────────────────
    for row_idx, diff_row in enumerate(body.diff_rows, start=3):
        status = diff_row.get("_diffStatus", "same")
        row_a  = diff_row.get("_rowA") or {}
        row_b  = diff_row.get("_rowB") or {}
        changed_fields = set(diff_row.get("_changedFields", []))

        status_label, bg_color = STATUS_MAP.get(status, ("＝相同", "FFFFFF"))
        row_fill   = PatternFill("solid", fgColor=bg_color) if bg_color != "FFFFFF" else None
        changed_fill = PatternFill("solid", fgColor=COLOR_CHANGED_CELL)

        # 状态列
        sc = ws.cell(row=row_idx, column=1, value=status_label)
        sc.alignment = center
        sc.border = thin
        if row_fill:
            sc.fill = row_fill
        sc.font = Font(size=10)

        for i, col in enumerate(cols):
            key = col.get("key", "")
            val_a = row_a.get(key, "（无）") if row_a else "（无）"
            val_b = row_b.get(key, "（无）") if row_b else "（无）"
            if isinstance(val_a, (dict, list)):
                val_a = json.dumps(val_a, ensure_ascii=False)
            if isinstance(val_b, (dict, list)):
                val_b = json.dumps(val_b, ensure_ascii=False)

            # A侧单元格
            ca = ws.cell(row=row_idx, column=2 + i, value=val_a if val_a != "（无）" else "")
            ca.border = thin
            ca.alignment = left_align
            ca.font = Font(size=10, bold=(status == "modified" and key in changed_fields))
            if status == "modified" and key in changed_fields:
                ca.fill = changed_fill
            elif row_fill:
                ca.fill = row_fill

            # B侧单元格
            cb = ws.cell(row=row_idx, column=n + 2 + i, value=val_b if val_b != "（无）" else "")
            cb.border = thin
            cb.alignment = left_align
            cb.font = Font(size=10, bold=(status == "modified" and key in changed_fields))
            if status == "modified" and key in changed_fields:
                cb.fill = changed_fill
            elif row_fill:
                cb.fill = row_fill

    # 冻结前两行
    ws.freeze_panes = "A3"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    file_b64 = base64.b64encode(buf.read()).decode()

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = body.filename or f"diff_report_{ts}.xlsx"
    return {"success": True, "data": {"file_b64": file_b64, "filename": filename}}


class DiffLarkSheetBody(BaseModel):
    user_access_token: str
    spreadsheet_token: str
    sheet_id: str = "Sheet1"
    columns:   List[dict]
    diff_rows: List[dict]
    label_a:   str = "数据集 A"
    label_b:   str = "数据集 B"


@router.post("/export/diff-lark-sheet")
async def export_diff_lark_sheet(
    body: DiffLarkSheetBody,
    request: Request,
    user=Depends(_READ),
    principal=Depends(get_authenticated_principal),
    gateway=Depends(get_default_gateway),
):
    data = await _invoke_data_exchange(
        request, user, principal, gateway,
        {
            "operation": "diff_lark_sheet", "user_access_token": body.user_access_token,
            "spreadsheet_token": body.spreadsheet_token, "sheet_id": body.sheet_id,
            "columns": body.columns, "diff_rows": body.diff_rows,
            "label_a": body.label_a, "label_b": body.label_b,
        },
    )
    return {"success": True, "data": data}


async def _legacy_export_diff_lark_sheet(body: DiffLarkSheetBody, user=Depends(_READ)):
    """将 diff 结果展开为平铺行写入飞书电子表格（含状态列）。"""
    STATUS_LABEL = {
        "added":    "+新增",
        "removed":  "-删除",
        "modified": "~变更",
        "same":     "=相同",
    }

    cols = body.columns
    # 构建表头：状态 + A侧各列（带标签前缀）+ B侧各列
    headers = ["状态"]
    for col in cols:
        headers.append(f"A-{col.get('label', col.get('key', ''))}")
    for col in cols:
        headers.append(f"B-{col.get('label', col.get('key', ''))}")

    rows = []
    for diff_row in body.diff_rows:
        status = diff_row.get("_diffStatus", "same")
        row_a  = diff_row.get("_rowA") or {}
        row_b  = diff_row.get("_rowB") or {}
        status_label = STATUS_LABEL.get(status, "=相同")

        flat_row = [status_label]
        for col in cols:
            key = col.get("key", "")
            v = row_a.get(key, "") if row_a else ""
            if isinstance(v, (dict, list)):
                v = json.dumps(v, ensure_ascii=False)
            flat_row.append(str(v) if v is not None else "")
        for col in cols:
            key = col.get("key", "")
            v = row_b.get(key, "") if row_b else ""
            if isinstance(v, (dict, list)):
                v = json.dumps(v, ensure_ascii=False)
            flat_row.append(str(v) if v is not None else "")
        rows.append(flat_row)

    values = [headers] + rows
    total_cols = len(headers)
    total_rows = len(values)

    url = (
        f"https://open.feishu.cn/open-apis/sheets/v2/spreadsheets"
        f"/{body.spreadsheet_token}/values"
    )
    payload = {
        "valueRange": {
            "range": f"{body.sheet_id}!A1:{_col_letter(total_cols)}{total_rows}",
            "values": values,
        }
    }
    async with httpx.AsyncClient(timeout=30) as client:
        resp = await client.put(
            url,
            headers={
                "Authorization": f"Bearer {body.user_access_token}",
                "Content-Type": "application/json",
            },
            json=payload,
        )
    if resp.status_code != 200:
        raise HTTPException(resp.status_code, f"飞书 API 错误：{resp.text}")
    data = resp.json()
    if data.get("code") != 0:
        raise HTTPException(400, f"飞书返回错误：{data.get('msg')}")
    return {"success": True, "data": {"written_rows": len(rows)}}
