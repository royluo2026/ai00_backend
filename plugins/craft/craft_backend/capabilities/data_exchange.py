"""Craft-owned bounded data-exchange export outcomes."""
from __future__ import annotations

import base64
import io
import json
from datetime import datetime
from typing import Any

import httpx

from backend.capability_v2.provider_contracts import (
    CapabilityBusinessError,
    CapabilityContext,
    CapabilityRisk,
    CapabilitySpec,
)



CAPABILITY_ID = "craft.data_exchange.export"
MAX_ROWS = 5000
MAX_COLUMNS = 200


def _json_cell(value: Any) -> Any:
    return json.dumps(value, ensure_ascii=False) if isinstance(value, (dict, list)) else value


def _save_workbook(workbook: Any) -> str:
    buf = io.BytesIO()
    workbook.save(buf)
    return base64.b64encode(buf.getvalue()).decode()


def _export_excel(payload: dict[str, Any], _context: CapabilityContext) -> dict[str, Any]:
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise CapabilityBusinessError("provider_unavailable", "openpyxl is required for Excel export") from exc

    config = payload.get("template_config") or {}
    rows = payload.get("rows") or []
    if not isinstance(config, dict) or not isinstance(rows, list):
        raise ValueError("template_config must be an object and rows must be an array")
    if len(rows) > MAX_ROWS:
        raise ValueError(f"rows must contain at most {MAX_ROWS} items")
    columns = [c for c in config.get("columns", []) if c.get("include", True)]
    if len(columns) > MAX_COLUMNS:
        raise ValueError(f"columns must contain at most {MAX_COLUMNS} items")
    styles = config.get("styles", {})
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "导出数据"
    header_bg = str(styles.get("headerBg", "#2563EB")).lstrip("#")
    header_fg = str(styles.get("headerFg", "#FFFFFF")).lstrip("#")
    alt_row_bg = str(styles.get("altRowBg", "#EFF6FF")).lstrip("#")
    font_size = int(styles.get("fontSize", 11))
    border_style = styles.get("borderStyle", "thin")
    thin_border = Border(*(Side(style=border_style) for _ in range(4)))
    center_align = Alignment(horizontal="center", vertical="center")
    for col_idx, col in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col.get("label", col.get("key", "")))
        cell.fill = PatternFill("solid", fgColor=header_bg)
        cell.font = Font(bold=True, color=header_fg, size=font_size)
        cell.border = thin_border
        cell.alignment = center_align
        ws.column_dimensions[get_column_letter(col_idx)].width = int(col.get("width", 15))
    ws.row_dimensions[1].height = 20
    alt_fill = PatternFill("solid", fgColor=alt_row_bg)
    for row_idx, row in enumerate(rows, 2):
        for col_idx, col in enumerate(columns, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=_json_cell(row.get(col["key"], "")))
            cell.border = thin_border
            if row_idx % 2 == 0:
                cell.fill = alt_fill
    filename = str(payload.get("filename") or f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
    return {"file_b64": _save_workbook(wb), "filename": filename}


def _export_diff_report(payload: dict[str, Any], _context: CapabilityContext) -> dict[str, Any]:
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise CapabilityBusinessError("provider_unavailable", "openpyxl is required for diff report export") from exc
    columns = payload.get("columns") or []
    diff_rows = payload.get("diff_rows") or []
    if not isinstance(columns, list) or not isinstance(diff_rows, list):
        raise ValueError("columns and diff_rows must be arrays")
    if len(columns) > MAX_COLUMNS or len(diff_rows) > MAX_ROWS:
        raise ValueError("diff report exceeds bounded response limits")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "差异对比"
    n = len(columns)
    thin = Border(*(Side(style="thin") for _ in range(4)))
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    colors = {"added": "DCFCE7", "removed": "FEE2E2", "modified": "FEF3C7", "same": "FFFFFF"}
    changed = PatternFill("solid", fgColor="F9E2AF")
    ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=n + 1)
    ws.merge_cells(start_row=1, start_column=n + 2, end_row=1, end_column=2 * n + 1)
    for col, value, color in ((2, payload.get("label_a") or "数据集 A", "1E3A5F"), (n + 2, payload.get("label_b") or "数据集 B", "1E4D2B")):
        cell = ws.cell(row=1, column=col, value=value)
        cell.fill = PatternFill("solid", fgColor=color)
        cell.font = Font(bold=True, color="FFFFFF", size=12)
        cell.alignment = center
    status = ws.cell(row=1, column=1, value="状态")
    status.fill = PatternFill("solid", fgColor="374151")
    status.font = Font(bold=True, color="FFFFFF", size=11)
    status.alignment = center
    ws.cell(row=2, column=1, value="").fill = PatternFill("solid", fgColor="374151")
    for i, col in enumerate(columns):
        label = col.get("label", col.get("key", ""))
        for offset, color in ((2, "1E3A5F"), (n + 2, "1E4D2B")):
            cell = ws.cell(row=2, column=offset + i, value=label)
            cell.fill = PatternFill("solid", fgColor=color)
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.alignment = center
            cell.border = thin
            ws.column_dimensions[get_column_letter(offset + i)].width = max(8, min(int(col.get("width", 12)), 30))
    ws.column_dimensions["A"].width = 8
    for row_idx, diff_row in enumerate(diff_rows, 3):
        kind = diff_row.get("_diffStatus", "same")
        row_fill = PatternFill("solid", fgColor=colors.get(kind, "FFFFFF"))
        row_a, row_b = diff_row.get("_rowA") or {}, diff_row.get("_rowB") or {}
        changed_fields = set(diff_row.get("_changedFields", []))
        cell = ws.cell(row=row_idx, column=1, value={"added": "+新增", "removed": "-删除", "modified": "~变更", "same": "=相同"}.get(kind, "=相同"))
        cell.alignment = center; cell.border = thin; cell.fill = row_fill
        for i, col in enumerate(columns):
            key = col.get("key", "")
            for offset, source in ((2, row_a), (n + 2, row_b)):
                cell = ws.cell(row=row_idx, column=offset + i, value=_json_cell(source.get(key, "")))
                cell.alignment = left; cell.border = thin
                cell.fill = changed if kind == "modified" and key in changed_fields else row_fill
    ws.freeze_panes = "A3"
    filename = str(payload.get("filename") or f"diff_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
    return {"file_b64": _save_workbook(wb), "filename": filename}


def _export_diff_lark_sheet(payload: dict[str, Any], _context: CapabilityContext) -> dict[str, Any]:
    token = str(payload.get("user_access_token") or "")
    spreadsheet = str(payload.get("spreadsheet_token") or "")
    if not token or not spreadsheet:
        raise ValueError("user_access_token and spreadsheet_token are required")
    columns = payload.get("columns") or []
    diff_rows = payload.get("diff_rows") or []
    headers = ["状态"] + [f"A-{c.get('label', c.get('key', ''))}" for c in columns] + [f"B-{c.get('label', c.get('key', ''))}" for c in columns]
    rows = []
    labels = {"added": "+新增", "removed": "-删除", "modified": "~变更", "same": "=相同"}
    for diff_row in diff_rows:
        status = diff_row.get("_diffStatus", "same")
        a, b = diff_row.get("_rowA") or {}, diff_row.get("_rowB") or {}
        rows.append([labels.get(status, "=相同")] + [str(_json_cell(a.get(c.get("key", ""), "")) or "") for c in columns] + [str(_json_cell(b.get(c.get("key", ""), "")) or "") for c in columns])
    def col_letter(n: int) -> str:
        result = ""
        while n:
            n, rem = divmod(n - 1, 26); result = chr(65 + rem) + result
        return result or "A"
    sheet_id = str(payload.get("sheet_id") or "Sheet1")
    values = [headers] + rows
    response = httpx.put(
        f"https://open.feishu.cn/open-apis/sheets/v2/spreadsheets/{spreadsheet}/values",
        headers={"Authorization": f"Bearer {token}", "Content-Type": "application/json"},
        json={"valueRange": {"range": f"{sheet_id}!A1:{col_letter(len(headers))}{len(values)}", "values": values}},
        timeout=30,
    )
    if response.status_code != 200:
        raise CapabilityBusinessError("external_service_error", f"Feishu API error: {response.text}")
    data = response.json()
    if data.get("code") != 0:
        raise CapabilityBusinessError("external_service_error", f"Feishu returned error: {data.get('msg')}")
    return {"written_rows": len(rows), "spreadsheet_token": spreadsheet, "sheet_id": sheet_id}


def export_data(payload: dict[str, Any], context: CapabilityContext) -> dict[str, Any]:
    operation = str(payload.get("operation") or "").strip()
    if operation == "excel":
        return _export_excel(payload, context)
    if operation == "diff_report":
        return _export_diff_report(payload, context)
    if operation == "diff_lark_sheet":
        return _export_diff_lark_sheet(payload, context)
    raise ValueError("operation must be one of: excel, diff_report, diff_lark_sheet")


def register_data_exchange_capability(registry: Any) -> None:
    registry.register(
        CapabilitySpec(
            id=CAPABILITY_ID,
            owner="craft",
            description="Export bounded Craft datasets to Excel or Feishu Sheets.",
            use_when="A governed Craft consumer needs a data-exchange export.",
            do_not_use_when="The operation mutates Craft business entities.",
            risk=CapabilityRisk.WRITE,
            confirmation="user",
            permissions=("craft.write",),
            idempotent=True,
            tags=("craft", "data-exchange", "export"),
        ),
        export_data,
    )


__all__ = ["CAPABILITY_ID", "export_data", "register_data_exchange_capability"]
